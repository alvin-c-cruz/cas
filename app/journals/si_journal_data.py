"""Pure data layer for the columnar Sales Invoice (SI) Journal.

No Flask request access here — callers pass plain dicts/values so these
functions are unit-testable in isolation.
"""
import io
from decimal import Decimal

from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# resolve_period is shared — import from the AP data layer, do not copy
from app.journals.ap_journal_data import resolve_period  # noqa: F401


def _column_sort_key_si(account, ar_account_id, wht_receivable_id, output_vat_ids):
    """Order: AR (0), Output VAT (1, by code), WHT Receivable (2), others (3, by code)."""
    if account.id == ar_account_id:
        return (0, account.code)
    if account.id in output_vat_ids:
        return (1, account.code)
    if account.id == wht_receivable_id:
        return (2, account.code)
    return (3, account.code)


def _group_si(account_id, ar_account_id, wht_receivable_id, output_vat_ids):
    if account_id == ar_account_id:
        return 'ar'
    if account_id == wht_receivable_id:
        return 'wht_recv'
    if account_id in output_vat_ids:
        return 'output_vat'
    return 'revenue'


def build_columnar_si(posted_entries, draft_entries, ar_account_id,
                      wht_receivable_id, output_vat_ids, voided_invoices=None):
    """Pivot journal-entry lines into a columnar SI matrix.

    Columns are built only from POSTED entries' accounts, ordered:
    AR (debit, 0), Output VAT (credit, 1), WHT Receivable (debit, 2),
    then other/revenue accounts (credit, 3).
    Posted rows carry signed amounts (debit - credit) per account and
    contribute to per-column totals. Draft and voided rows are listed
    with a flag and no amounts, excluded from totals.

    Returns dict: columns, rows, totals, grand_total, balanced.
    """
    if voided_invoices is None:
        voided_invoices = []
    if output_vat_ids is None:
        output_vat_ids = set()

    accounts_by_id = {}
    totals = {}
    rows = []

    for je in posted_entries:
        cells = {}
        for line in je.lines.all():
            acct = line.account
            accounts_by_id[acct.id] = acct
            signed = (line.debit_amount or Decimal('0')) - (line.credit_amount or Decimal('0'))
            cells[acct.id] = cells.get(acct.id, Decimal('0')) + signed
            totals[acct.id] = totals.get(acct.id, Decimal('0')) + signed
        rows.append({'entry': je, 'cells': cells, 'is_draft': False, 'is_voided': False})

    for je in draft_entries:
        rows.append({'entry': je, 'cells': {}, 'is_draft': True, 'is_voided': False})

    for invoice in voided_invoices:
        rows.append({'invoice': invoice, 'entry': None, 'cells': {}, 'is_draft': False, 'is_voided': True})

    ordered = sorted(
        accounts_by_id.values(),
        key=lambda a: _column_sort_key_si(a, ar_account_id, wht_receivable_id, output_vat_ids),
    )
    columns = [{
        'account_id': a.id,
        'code': a.code,
        'name': a.name,
        'group': _group_si(a.id, ar_account_id, wht_receivable_id, output_vat_ids),
    } for a in ordered]

    def _row_sort_key(r):
        if r['is_voided']:
            return (r['invoice'].invoice_date, r['invoice'].invoice_number)
        return (r['entry'].entry_date, r['entry'].entry_number or '')

    rows.sort(key=_row_sort_key)

    grand_total = sum(totals.values(), Decimal('0'))
    return {
        'columns': columns,
        'rows': rows,
        'totals': totals,
        'grand_total': grand_total,
        'balanced': grand_total == Decimal('0'),
    }


def _fmt(value):
    """Render a signed Decimal: credits (negative) in parentheses, blanks for zero/None."""
    if value is None or value == Decimal('0'):
        return ''
    if value < 0:
        return f'({-value:,.2f})'
    return f'{value:,.2f}'


def build_si_journal_xlsx(columns, rows, totals, period_label, company_name,
                          branch_name, filename, identity):
    """Build the columnar SI Journal as an .xlsx Flask response.

    branch_name=None skips the branch row (caller passes None when only one
    branch exists). Amount columns use SUM formulas so totals are live in Excel.

    identity(entry) must return (si_no, customer_po, customer_name, notes).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'SI Journal'

    right = Alignment(horizontal='right')
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    num_fmt = '#,##0.00;(#,##0.00)'

    thin = Side(style='thin')
    double_s = Side(style='double')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_border = Border(bottom=double_s)
    draft_fill  = PatternFill(fill_type='solid', fgColor='FFF9C4')  # light yellow
    voided_fill = PatternFill(fill_type='solid', fgColor='FFCDD2')  # light red

    # Preamble
    ws.append([company_name])
    ws['A1'].font = Font(bold=True, size=16)

    if branch_name:
        ws.append([branch_name])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=16)

    ws.append(['Sales Journal'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14)
    ws.append([period_label])
    ws.append([])

    # Header row
    fixed = ['Date', 'SI No.', 'Customer PO', 'Customer', 'Particulars']
    header = fixed + [c['name'] for c in columns]
    ws.append(header)
    hdr_row = ws.max_row
    ws.row_dimensions[hdr_row].height = 40
    for cell in ws[hdr_row]:
        cell.font = Font(bold=True)
        cell.alignment = center_wrap
        cell.border = cell_border

    # Data rows
    first_data_row = hdr_row + 1
    for r in rows:
        if r.get('is_voided'):
            inv = r['invoice']
            line = [
                inv.invoice_date.strftime('%d-%b-%Y'),
                inv.invoice_number or '',
                getattr(inv, 'customer_po_number', '') or '',
                inv.customer_name or '',
                '[VOIDED] ' + (inv.notes or ''),
            ] + [None] * len(columns)
            ws.append(line)
            cur = ws.max_row
            for cell in ws[cur]:
                cell.border = cell_border
                cell.fill = voided_fill
            continue

        e = r['entry']
        si_no, customer_po, customer_name, notes = identity(e)
        line = [
            e.entry_date.strftime('%d-%b-%Y'),
            si_no or '',
            customer_po or '',
            customer_name or '',
            ('[DRAFT] ' + (notes or '')) if r['is_draft'] else (notes or ''),
        ]
        for c in columns:
            if r['is_draft']:
                line.append(None)
            else:
                val = r['cells'].get(c['account_id'])
                line.append(float(val) if val is not None else None)
        ws.append(line)
        cur = ws.max_row
        for i, cell in enumerate(ws[cur], 1):
            cell.border = cell_border
            if r['is_draft']:
                cell.fill = draft_fill
            if i > len(fixed):
                cell.number_format = num_fmt
                cell.alignment = right

    last_data_row = ws.max_row

    # Blank separator before total
    ws.append([])

    # TOTAL row: SUM formulas, double bottom rule only
    ws.append(['TOTAL', '', '', '', ''])
    tot_row = ws.max_row
    for i in range(1, len(fixed) + 1):
        ws.cell(row=tot_row, column=i).font = Font(bold=True)
        ws.cell(row=tot_row, column=i).border = total_border
    for i, c in enumerate(columns, len(fixed) + 1):
        col_letter = get_column_letter(i)
        cell = ws.cell(row=tot_row, column=i)
        cell.value = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        cell.font = Font(bold=True)
        cell.number_format = num_fmt
        cell.alignment = right
        cell.border = total_border

    # Column widths
    col_widths = [12, 22, 22, 28, 40] + [20] * len(columns)
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
