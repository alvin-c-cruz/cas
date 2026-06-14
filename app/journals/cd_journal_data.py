"""Pure data layer for the columnar Cash Disbursements Journal.

No Flask request access here — callers pass plain dicts/values so these
functions are unit-testable in isolation.
"""
import io
from decimal import Decimal

from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _group(account, ap_account_id, wt_account_id, input_vat_account_ids, totals):
    """Classify an account into a CD journal column group.

    Groups determine column colour (debit=blue, credit=red) and sort order.
    Cash/bank accounts are identified by their net credit sign in totals.
    """
    if account.id == ap_account_id:
        return 'ap_applied'  # debit — paying down AP
    if account.id == wt_account_id:
        return 'wht'          # credit
    if account.id in input_vat_account_ids:
        return 'vat'          # debit
    if totals.get(account.id, Decimal('0')) < 0:
        return 'cash'         # credit — net outflow
    return 'expense'          # debit


def _col_key(account, ap_account_id, wt_account_id, input_vat_account_ids, totals):
    """Sort: ap_applied(0) -> vat(1) -> expense(2) -> wht(3) -> cash(4), then by code."""
    order = {'ap_applied': 0, 'vat': 1, 'expense': 2, 'wht': 3, 'cash': 4}
    g = _group(account, ap_account_id, wt_account_id, input_vat_account_ids, totals)
    return (order.get(g, 2), account.code)


def build_columnar_cd(posted_entries, draft_entries, ap_account_id,
                      wt_account_id, input_vat_account_ids, cancelled_refs=None):
    """Pivot disbursement journal-entry lines into a columnar matrix.

    Columns are built only from POSTED entries' accounts. Cancelled CDVs
    (whose JEs remain posted) are included in totals but flagged is_cancelled=True
    for visual display (strikethrough + badge). Draft rows carry no amounts.

    cancelled_refs: set of CDV numbers (entry.reference) that are cancelled.

    Returns dict: columns, rows, totals, grand_total, balanced.
    """
    if cancelled_refs is None:
        cancelled_refs = set()

    accounts_by_id = {}
    totals = {}
    rows = []

    for je in posted_entries:
        cells = {}
        for line in je.lines.all():
            acct = line.account
            accounts_by_id[acct.id] = acct
            signed = (line.debit_amount or Decimal('0')) - (line.credit_amount or Decimal('0'))
            cells[acct.id] = cells.get(acct.id, Decimal('0')) + signed
            totals[acct.id] = totals.get(acct.id, Decimal('0')) + signed
        is_cancelled = je.reference in cancelled_refs if cancelled_refs else False
        rows.append({'entry': je, 'cells': cells, 'is_draft': False, 'is_cancelled': is_cancelled})

    for je in draft_entries:
        rows.append({'entry': je, 'cells': {}, 'is_draft': True, 'is_cancelled': False})

    ordered = sorted(
        accounts_by_id.values(),
        key=lambda a: _col_key(a, ap_account_id, wt_account_id, input_vat_account_ids, totals),
    )
    columns = [{
        'account_id': a.id,
        'code': a.code,
        'name': a.name,
        'group': _group(a, ap_account_id, wt_account_id, input_vat_account_ids, totals),
    } for a in ordered]

    rows.sort(key=lambda r: (r['entry'].entry_date, r['entry'].entry_number))

    grand_total = sum(totals.values(), Decimal('0'))
    return {
        'columns': columns,
        'rows': rows,
        'totals': totals,
        'grand_total': grand_total,
        'balanced': grand_total == Decimal('0'),
    }


def _fmt(value):
    """Render a signed Decimal: credits (negative) in parentheses, blanks for zero/None."""
    if value is None or value == Decimal('0'):
        return ''
    if value < 0:
        return f'({-value:,.2f})'
    return f'{value:,.2f}'


def build_cd_journal_xlsx(columns, rows, totals, period_label, company_name,
                          branch_name, filename, identity):
    """Build the columnar CD Journal as an .xlsx Flask response.

    branch_name=None skips the branch row.
    identity(entry) -> (cd_no, check_no, vendor, particulars)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'CD Journal'

    right = Alignment(horizontal='right')
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    num_fmt = '#,##0.00;(#,##0.00)'

    thin = Side(style='thin')
    double_s = Side(style='double')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_border = Border(bottom=double_s)
    draft_fill     = PatternFill(fill_type='solid', fgColor='FFF9C4')  # light yellow
    cancelled_fill = PatternFill(fill_type='solid', fgColor='FFCDD2')  # light red

    ws.append([company_name])
    ws['A1'].font = Font(bold=True, size=16)
    if branch_name:
        ws.append([branch_name])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=16)
    ws.append(['Cash Disbursements Journal'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14)
    ws.append([period_label])
    ws.append([])

    fixed = ['Date', 'CD No.', 'Check No.', 'Vendor', 'Particulars']
    header = fixed + [c['name'] for c in columns]
    ws.append(header)
    hdr_row = ws.max_row
    ws.row_dimensions[hdr_row].height = 40
    for cell in ws[hdr_row]:
        cell.font = Font(bold=True)
        cell.alignment = center_wrap
        cell.border = cell_border

    first_data_row = hdr_row + 1
    for r in rows:
        e = r['entry']
        no, check_no, vendor, particulars = identity(e)
        if r['is_cancelled']:
            particulars = '[CANCELLED] ' + (particulars or '')
        elif r['is_draft']:
            particulars = '[DRAFT] ' + (particulars or '')
        line = [
            e.entry_date.strftime('%d-%b-%Y'),
            no or '',
            check_no or '',
            vendor or '',
            particulars or '',
        ]
        for c in columns:
            if r['is_draft']:
                line.append(None)
            else:
                val = r['cells'].get(c['account_id'])
                line.append(float(val) if val else None)
        ws.append(line)
        cur = ws.max_row
        fill = cancelled_fill if r['is_cancelled'] else (draft_fill if r['is_draft'] else None)
        for i, cell in enumerate(ws[cur], 1):
            cell.border = cell_border
            if fill:
                cell.fill = fill
            if i > len(fixed):
                cell.number_format = num_fmt
                cell.alignment = right

    last_data_row = ws.max_row
    ws.append([])
    ws.append(['TOTAL', '', '', '', ''])
    tot_row = ws.max_row
    for i in range(1, len(fixed) + 1):
        ws.cell(row=tot_row, column=i).font = Font(bold=True)
        ws.cell(row=tot_row, column=i).border = total_border
    for i, c in enumerate(columns, len(fixed) + 1):
        col_letter = get_column_letter(i)
        cell = ws.cell(row=tot_row, column=i)
        cell.value = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row + 1})'
        cell.font = Font(bold=True)
        cell.number_format = num_fmt
        cell.alignment = right
        cell.border = total_border

    col_widths = [12, 22, 18, 28, 40] + [20] * len(columns)
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
