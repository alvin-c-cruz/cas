"""Shared helpers for BIR books of accounts: company-identity lookup and the
Excel header block written at the top of every exported book sheet."""
from app.settings import AppSettings


def get_company_identity():
    g = AppSettings.get_setting
    return {
        'name': g('company_name') or '',
        'tin': g('company_tin') or '',
        'tin_branch': g('tin_branch_code') or '',
        'rdo': g('rdo_code') or '',
        'address': g('company_address') or '',
        'sss_employer_no': g('sss_employer_no') or '',
        'philhealth_employer_no': g('philhealth_employer_no') or '',
        'pagibig_employer_no': g('pagibig_employer_no') or '',
    }


def tin_line(company):
    parts = []
    if company.get('tin'):
        tin = company['tin']
        if company.get('tin_branch'):
            tin += '-' + company['tin_branch']
        parts.append('TIN: ' + tin)
    if company.get('rdo'):
        parts.append('RDO: ' + company['rdo'])
    return ' · '.join(parts)


def write_bir_book_header(ws, company, book_title, period_label, branch_name=None):
    from openpyxl.styles import Font

    def put(text='', bold=False, size=11):
        ws.append([text])
        if bold or size != 11:
            ws.cell(ws.max_row, 1).font = Font(bold=bold, size=size)
        return ws.max_row

    if company.get('name'):
        put(company['name'].upper(), bold=True, size=14)
    line = tin_line(company)
    if line:
        put(line)
    if company.get('address'):
        put(company['address'])
    if branch_name:
        put('Branch: ' + branch_name)
    put(book_title, bold=True, size=13)
    put(period_label)
    put()
    return ws.max_row
