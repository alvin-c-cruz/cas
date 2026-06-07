"""
Export utilities for Excel and CSV formats.

Provides functions to export lists and reports to:
- Excel (.xlsx) - formatted with headers
- CSV - simple comma-separated values

Usage:
    from app.utils.export import export_to_excel, export_to_csv

    # Export invoices
    response = export_to_excel(
        data=invoices,
        columns=['invoice_number', 'invoice_date', 'customer_name', 'total_amount'],
        headers=['Invoice #', 'Date', 'Customer', 'Total'],
        filename='sales_invoices.xlsx'
    )
"""
import csv
import io
from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, datetime
from decimal import Decimal


def format_value(value):
    """Format value for export (convert dates, decimals, etc.)"""
    if value is None:
        return ''
    elif isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d') if isinstance(value, date) else value.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(value, Decimal):
        return float(value)
    elif isinstance(value, bool):
        return 'Yes' if value else 'No'
    else:
        return str(value)


def get_nested_value(obj, field):
    """Get value from object, supports nested attributes (e.g., 'customer.name')"""
    if '.' in field:
        parts = field.split('.')
        value = obj
        for part in parts:
            if hasattr(value, part):
                value = getattr(value, part)
            else:
                return None
        return value
    else:
        return getattr(obj, field, None)


def export_to_excel(data, columns, headers, filename='export.xlsx', title=None):
    """
    Export data to Excel format with formatting.

    Args:
        data: List of objects (SQLAlchemy models, dicts, etc.)
        columns: List of column names/attributes to export
        headers: List of header labels for display
        filename: Output filename (default: export.xlsx)
        title: Optional title row

    Returns:
        Flask Response object with Excel file
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_num = 1

    # Add title row if provided
    if title:
        ws.merge_cells(f'A1:{chr(64 + len(headers))}1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        row_num = 2

    # Add headers
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

        # Auto-adjust column width
        ws.column_dimensions[chr(64 + col_num)].width = max(len(str(header)) + 2, 12)

    row_num += 1

    # Add data rows
    for item in data:
        for col_num, column in enumerate(columns, start=1):
            # Get value from object
            if isinstance(item, dict):
                value = item.get(column)
            else:
                value = get_nested_value(item, column)

            # Format and write value
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = format_value(value)
            cell.border = border

            # Format numbers with alignment
            if isinstance(value, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(value, Decimal) or isinstance(value, float):
                    cell.number_format = '#,##0.00'

        row_num += 1

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


def export_to_csv(data, columns, headers, filename='export.csv'):
    """
    Export data to CSV format.

    Args:
        data: List of objects (SQLAlchemy models, dicts, etc.)
        columns: List of column names/attributes to export
        headers: List of header labels for display
        filename: Output filename (default: export.csv)

    Returns:
        Flask Response object with CSV file
    """
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Write headers
    writer.writerow(headers)

    # Write data rows
    for item in data:
        row = []
        for column in columns:
            # Get value from object
            if isinstance(item, dict):
                value = item.get(column)
            else:
                value = get_nested_value(item, column)

            # Format and add to row
            row.append(format_value(value))

        writer.writerow(row)

    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


def export_dict_to_excel(data_dict, filename='export.xlsx', title=None):
    """
    Export dictionary data to Excel (useful for reports with multiple sections).

    Args:
        data_dict: Dict with keys as section names, values as list of dicts
        filename: Output filename
        title: Optional title

    Example:
        data = {
            'Summary': [
                {'Metric': 'Total Sales', 'Value': 1000000},
                {'Metric': 'Total Purchases', 'Value': 750000}
            ],
            'Details': [...]
        }
    """
    wb = Workbook()

    for sheet_name, rows in data_dict.items():
        if not rows:
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit: 31 chars

        # Get headers from first row keys
        headers = list(rows[0].keys()) if rows else []

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Write data
        for row_num, row_data in enumerate(rows, start=2):
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = format_value(row_data.get(header))

    # Remove default sheet if we created others
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb['Sheet'])

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response
