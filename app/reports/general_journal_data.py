"""Classic two-column General Journal (book of original entry) built from
voucher-type journal entries. Lives in reports/ so the journals blueprint stays
untouched; reads the JournalEntry model only."""
from decimal import Decimal
from io import BytesIO

VOUCHER_ENTRY_TYPES = ('reversal', 'adjustment', 'closing', 'closing_reversal',
                       'opening', 'opening_balance', 'reclassification', 'transfer',
                       'petty_cash_replenishment')


def build_general_journal(entries):
    """Shape an iterable of JournalEntry into General Journal rows. Only posted
    entries contribute to totals; drafts and voided entries are listed but excluded."""
    rows, total_debit, total_credit = [], Decimal('0.00'), Decimal('0.00')
    for e in entries:
        debits, credits = [], []
        for line in e.lines:
            if line.debit_amount and line.debit_amount > 0:
                debits.append({'account': line.account, 'amount': line.debit_amount})
            if line.credit_amount and line.credit_amount > 0:
                credits.append({'account': line.account, 'amount': line.credit_amount})
        if e.status == 'posted':
            total_debit += sum((d['amount'] for d in debits), Decimal('0.00'))
            total_credit += sum((c['amount'] for c in credits), Decimal('0.00'))
        rows.append({'entry': e, 'debits': debits, 'credits': credits,
                     'explanation': e.description or '',
                     'is_draft': e.status == 'draft',
                     'is_voided': e.status in ('cancelled', 'reversed')})
    return {'rows': rows, 'total_debit': total_debit, 'total_credit': total_credit,
            'balanced': total_debit == total_credit}


def _write_gj_rows(ws, gj):
    """Write GJ data rows + TOTAL row to ws. Callers apply number formats and widths."""
    from openpyxl.styles import Font
    for row in gj['rows']:
        e = row['entry']
        ws.append([e.entry_date.strftime('%m/%d/%Y'), '', e.display_number, None, None])
        for d in row['debits']:
            ws.append(['', d['account'].name, '', float(d['amount']), None])
        for c in row['credits']:
            ws.append(['', '    ' + c['account'].name, '', None, float(c['amount'])])
        if row['explanation']:
            ws.append(['', '(' + row['explanation'] + ')', '', None, None])
    ws.append(['', 'TOTAL', '', float(gj['total_debit']), float(gj['total_credit'])])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def build_general_journal_xlsx(gj, period_label, company, branch_name, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from app.utils.bir_books import write_bir_book_header

    NUM = '#,##0.00;(#,##0.00)'
    wb = Workbook(); ws = wb.active; ws.title = 'General Journal'
    write_bir_book_header(ws, company, 'GENERAL JOURNAL', period_label, branch_name)
    ws.append(['Date', 'Particulars', 'Ref', 'Debit', 'Credit'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    _write_gj_rows(ws, gj)
    for col in ('D', 'E'):
        for cell in ws[col]:
            cell.number_format = NUM
    for col, w in {'A': 12, 'B': 50, 'C': 16, 'D': 16, 'E': 16}.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return bio
