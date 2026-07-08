"""
Reports views for financial reporting.
Includes AR Aging, AP Aging, BIR compliance reports, and financial statements.
"""
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, session
from flask_login import login_required, current_user
from functools import wraps
from app import db
from app.sales_invoices.models import SalesInvoice
from app.accounts.models import Account
from app.accounts_payable.models import AccountsPayable
from app.cash_receipts.models import CashReceiptVoucher
from app.cash_disbursements.models import CashDisbursementVoucher
from app.reports.bir import (
    get_summary_list_of_sales,
    get_summary_list_of_purchases,
    get_alphalist_of_payees,
    get_month_name,
    get_quarter_name,
    get_quarter_months
)
from app.reports.financial import (
    generate_trial_balance,
    generate_income_statement,
    generate_balance_sheet,
    generate_general_ledger,
    generate_cash_flow,
)
from app.utils.export import export_to_excel, export_to_csv
from app.utils.bir_books import get_company_identity
from app.journal_entries.models import JournalEntry
from app.reports.general_journal_data import (build_general_journal,
    build_general_journal_xlsx, _write_gj_rows, VOUCHER_ENTRY_TYPES)
from app.reports.books_data import BOOKS
from app.reports.two_column import merge_is_two_column
from app.journals.ap_journal_data import resolve_period   # pure fn, not modified
from app.utils import ph_now, end_of_month
from datetime import date, timedelta, datetime
from decimal import Decimal
from sqlalchemy import func

reports_bp = Blueprint('reports', __name__, template_folder='templates')

# entry_type -> (Model, number column, view endpoint, short label prefix)
_SOURCE_MAP = {
    'sale':         (SalesInvoice,            'invoice_number', 'sales_invoices.view',    'SI'),
    'purchase':     (AccountsPayable,         'ap_number',      'accounts_payable.view',  'AP'),
    'receipt':      (CashReceiptVoucher,      'crv_number',     'cash_receipts.view',     'CR'),
    'disbursement': (CashDisbursementVoucher, 'cdv_number',     'cash_disbursements.view', 'CD'),
}


def _attach_source_links(ledger, branch_id):
    """Mutate each line, adding line['source'] = {'url', 'label'}.

    Resolves the four auto-posted transaction types to their source document by
    number (one IN-query per type); everything else links to the Journal Entry view.
    """
    # Gather the distinct references actually present, grouped by entry_type.
    refs_by_type = {}
    for account in ledger['accounts']:
        for line in account['lines']:
            et = line.get('entry_type')
            if et in _SOURCE_MAP and line.get('reference'):
                refs_by_type.setdefault(et, set()).add(line['reference'])

    # Build {number: id} maps with one query per source type.
    id_maps = {}
    for et, refs in refs_by_type.items():
        model, numcol, _endpoint, _prefix = _SOURCE_MAP[et]
        col = getattr(model, numcol)
        rows = model.query.filter(model.branch_id == branch_id, col.in_(refs)).all()
        id_maps[et] = {getattr(r, numcol): r.id for r in rows}

    for account in ledger['accounts']:
        for line in account['lines']:
            et = line.get('entry_type')
            ref = line.get('reference')
            mapped = _SOURCE_MAP.get(et)
            doc_id = id_maps.get(et, {}).get(ref) if mapped else None
            if mapped and doc_id is not None:
                _model, _numcol, endpoint, prefix = mapped
                line['source'] = {'url': url_for(endpoint, id=doc_id),
                                  'label': f'{prefix} {ref}'}
            else:
                line['source'] = {'url': url_for('journal_entries.view', id=line['entry_id']),
                                  'label': line.get('display_number') or line['entry_number']}


def accountant_or_admin_required(f):
    """Decorator to require accountant or admin role."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('users.login'))
        if not (current_user.role == 'accountant' or current_user.has_full_access):
            flash('Only Accountants and Administrators can access reports.', 'error')
            return redirect(url_for('dashboard.index'))
        return f(*args, **kwargs)
    return decorated_function


def calculate_age_bucket(due_date, as_of_date):
    """
    Calculate which age bucket a date falls into.
    Returns: 'current', '1-30', '31-60', '61-90', '90+'
    """
    if not due_date:
        return 'current'

    days_overdue = (as_of_date - due_date).days

    if days_overdue <= 0:
        return 'current'
    elif days_overdue <= 30:
        return '1-30'
    elif days_overdue <= 60:
        return '31-60'
    elif days_overdue <= 90:
        return '61-90'
    else:
        return '90+'


def _build_ar_aging_data(as_of_date, branch_id):
    """Build AR aging data for the given as_of_date and branch.

    Returns (customers_list, grand_totals).
    customers_list: list of dicts, each:
      {'name': str, 'invoices': [...], 'current': Decimal, '1-30': Decimal,
       '31-60': Decimal, '61-90': Decimal, '90+': Decimal, 'total': Decimal}
      sorted by total desc.
    grand_totals: dict with keys 'current','1-30','31-60','61-90','90+','total' as Decimals.
    """
    invoices = SalesInvoice.query.filter(
        SalesInvoice.status.in_(['posted', 'partially_paid']),
        SalesInvoice.balance > 0,
        SalesInvoice.branch_id == branch_id
    ).order_by(SalesInvoice.customer_name, SalesInvoice.due_date).all()

    customers = {}
    for invoice in invoices:
        key = invoice.customer_name
        if key not in customers:
            customers[key] = {
                'name': invoice.customer_name,
                'invoices': [],
                'current': Decimal('0.00'),
                '1-30': Decimal('0.00'),
                '31-60': Decimal('0.00'),
                '61-90': Decimal('0.00'),
                '90+': Decimal('0.00'),
                'total': Decimal('0.00'),
            }
        bucket = calculate_age_bucket(invoice.due_date, as_of_date)
        customers[key]['invoices'].append({
            'invoice_id': invoice.id,
            'invoice_number': invoice.invoice_number,
            'invoice_date': invoice.invoice_date,
            'due_date': invoice.due_date,
            'balance_due': invoice.balance,
            'bucket': bucket,
            'days_overdue': max(0, (as_of_date - invoice.due_date).days) if invoice.due_date else 0,
        })
        customers[key][bucket] += invoice.balance
        customers[key]['total'] += invoice.balance

    grand_totals = {
        'current': Decimal('0.00'), '1-30': Decimal('0.00'),
        '31-60': Decimal('0.00'), '61-90': Decimal('0.00'),
        '90+': Decimal('0.00'), 'total': Decimal('0.00'),
    }
    for c in customers.values():
        for k in grand_totals:
            grand_totals[k] += c[k]

    customers_list = sorted(customers.values(), key=lambda x: x['total'], reverse=True)
    return customers_list, grand_totals


def _build_ap_aging_data(as_of_date, branch_id):
    """Build AP aging data for the given as_of_date and branch.

    Returns (vendors_list, grand_totals).
    vendors_list: list of dicts, each:
      {'name': str, 'bills': [...], 'current': Decimal, '1-30': Decimal,
       '31-60': Decimal, '61-90': Decimal, '90+': Decimal, 'total': Decimal}
      sorted by total desc.
    grand_totals: dict with keys 'current','1-30','31-60','61-90','90+','total' as Decimals.
    """
    bills = AccountsPayable.query.filter(
        AccountsPayable.status.in_(['posted', 'partially_paid']),
        AccountsPayable.balance > 0,
        AccountsPayable.branch_id == branch_id,
        AccountsPayable.payee_type == 'vendor',   # exclude employee-payee vouchers
    ).order_by(AccountsPayable.vendor_name, AccountsPayable.due_date).all()

    vendors = {}
    for bill in bills:
        key = bill.vendor_name
        if key not in vendors:
            vendors[key] = {
                'name': bill.vendor_name,
                'bills': [],
                'current': Decimal('0.00'),
                '1-30': Decimal('0.00'),
                '31-60': Decimal('0.00'),
                '61-90': Decimal('0.00'),
                '90+': Decimal('0.00'),
                'total': Decimal('0.00'),
            }
        bucket = calculate_age_bucket(bill.due_date, as_of_date)
        vendors[key]['bills'].append({
            'ap_id': bill.id,
            'ap_number': bill.ap_number,
            'ap_date': bill.ap_date,
            'due_date': bill.due_date,
            'balance_due': bill.balance,
            'bucket': bucket,
            'days_overdue': max(0, (as_of_date - bill.due_date).days) if bill.due_date else 0,
        })
        vendors[key][bucket] += bill.balance
        vendors[key]['total'] += bill.balance

    grand_totals = {
        'current': Decimal('0.00'), '1-30': Decimal('0.00'),
        '31-60': Decimal('0.00'), '61-90': Decimal('0.00'),
        '90+': Decimal('0.00'), 'total': Decimal('0.00'),
    }
    for v in vendors.values():
        for k in grand_totals:
            grand_totals[k] += v[k]

    vendors_list = sorted(vendors.values(), key=lambda x: x['total'], reverse=True)
    return vendors_list, grand_totals


@reports_bp.route('/reports')
@login_required
def index():
    """Reports dashboard."""
    return render_template('reports/index.html')


@reports_bp.route('/reports/ar-aging')
@login_required
def ar_aging():
    as_of_str = request.args.get('as_of', date.today().isoformat())
    try:
        as_of_date = date.fromisoformat(as_of_str)
    except ValueError:
        as_of_date = date.today()

    current_branch_id = session.get('selected_branch_id')
    customers_list, grand_totals = _build_ar_aging_data(as_of_date, current_branch_id)
    return render_template('reports/ar_aging.html',
                           customers=customers_list,
                           grand_totals=grand_totals,
                           as_of_date=as_of_date)


@reports_bp.route('/reports/ap-aging')
@login_required
def ap_aging():
    as_of_str = request.args.get('as_of', date.today().isoformat())
    try:
        as_of_date = date.fromisoformat(as_of_str)
    except ValueError:
        as_of_date = date.today()

    current_branch_id = session.get('selected_branch_id')
    vendors_list, grand_totals = _build_ap_aging_data(as_of_date, current_branch_id)
    return render_template('reports/ap_aging.html',
                           vendors=vendors_list,
                           grand_totals=grand_totals,
                           as_of_date=as_of_date)


def _gl_params():
    """Shared (start_date, end_date, account_id, branch_id) parsing for GL routes."""
    today = date.today()
    start_default = date(today.year, 1, 1)  # first day of the current year

    def _parse(param, fallback):
        try:
            return date.fromisoformat(request.args.get(param, ''))
        except (ValueError, TypeError):
            return fallback

    return (_parse('start_date', start_default),
            _parse('end_date', today),
            request.args.get('account_id', type=int),
            session.get('selected_branch_id'))


def _flatten_ledger(ledger):
    """Flatten the book into export rows: a header row per account, its lines, a subtotal."""
    rows = []
    for acct in ledger['accounts']:
        rows.append({'date': f"{acct['code']} - {acct['name']}", 'source': '',
                     'particulars': 'Opening balance', 'debit': '', 'credit': '',
                     'balance': acct['opening_balance']})
        for line in acct['lines']:
            rows.append({
                'date': line['entry_date'],
                'source': line['source']['label'], 'particulars': line['description'],
                'debit': line['debit'] or '', 'credit': line['credit'] or '',
                'balance': line['running_balance'],
            })
        rows.append({'date': '', 'source': '', 'particulars': 'Closing balance',
                     'debit': acct['total_debit'], 'credit': acct['total_credit'],
                     'balance': acct['closing_balance']})
    return rows


_GL_COLUMNS = ['date', 'source', 'particulars', 'debit', 'credit', 'balance']
_GL_HEADERS = ['Date', 'Source', 'Particulars', 'Debit', 'Credit', 'Balance']


@reports_bp.route('/reports/general-ledger')
@login_required
def general_ledger():
    """All-accounts General Ledger book for the selected branch."""
    start_date, end_date, account_id, branch_id = _gl_params()
    ledger = generate_general_ledger(start_date, end_date, branch_id, account_id=account_id)
    _attach_source_links(ledger, branch_id)

    accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    return render_template('reports/general_ledger.html',
                           ledger=ledger,
                           start_date=start_date,
                           end_date=end_date,
                           accounts=accounts,
                           selected_account_id=account_id)


@reports_bp.route('/reports/general-ledger/export/excel')
@login_required
def general_ledger_export_excel():
    start_date, end_date, account_id, branch_id = _gl_params()
    ledger = generate_general_ledger(start_date, end_date, branch_id, account_id=account_id)
    _attach_source_links(ledger, branch_id)
    rows = _flatten_ledger(ledger)
    return export_to_excel(
        rows, _GL_COLUMNS, _GL_HEADERS,
        filename=f'general_ledger_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx',
        title=f'General Ledger - {start_date.isoformat()} to {end_date.isoformat()}')


@reports_bp.route('/reports/general-ledger/export/csv')
@login_required
def general_ledger_export_csv():
    start_date, end_date, account_id, branch_id = _gl_params()
    ledger = generate_general_ledger(start_date, end_date, branch_id, account_id=account_id)
    _attach_source_links(ledger, branch_id)
    rows = _flatten_ledger(ledger)
    return export_to_csv(
        rows, _GL_COLUMNS, _GL_HEADERS,
        filename=f'general_ledger_{start_date.isoformat()}_to_{end_date.isoformat()}.csv')


@reports_bp.route('/reports/general-ledger/print')
@login_required
def general_ledger_print():
    start_date, end_date, account_id, branch_id = _gl_params()
    ledger = generate_general_ledger(start_date, end_date, branch_id, account_id=account_id)
    _attach_source_links(ledger, branch_id)
    period_label = (f"{start_date.strftime('%b %d, %Y')} to "
                    f"{end_date.strftime('%b %d, %Y')}")
    return render_template('reports/general_ledger_print.html',
                           ledger=ledger, start_date=start_date, end_date=end_date,
                           company=get_company_identity(), period_label=period_label)


# BIR Compliance Reports

@reports_bp.route('/reports/bir')
@login_required
@accountant_or_admin_required
def bir_index():
    return redirect(url_for('dashboard.under_development', feature='BIR Reports'))
    current_year = datetime.now().year
    current_month = datetime.now().month
    current_quarter = ((current_month - 1) // 3) + 1

    return render_template('reports/bir_index.html',
                         current_year=current_year,
                         current_month=current_month,
                         current_quarter=current_quarter)


@reports_bp.route('/reports/bir/sales')
@login_required
@accountant_or_admin_required
def bir_sales():
    """Summary List of Sales (Annex A) - Monthly VAT Sales"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    current_branch_id = session.get('selected_branch_id')
    sales_data = get_summary_list_of_sales(year, month, branch_id=current_branch_id)

    return render_template('reports/bir_sales.html',
                         sales_data=sales_data,
                         year=year,
                         month=month,
                         month_name=get_month_name(month),
                         company=get_company_identity())


@reports_bp.route('/reports/bir/sales/export/excel')
@login_required
@accountant_or_admin_required
def bir_sales_export_excel():
    """Export Summary List of Sales to Excel"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    current_branch_id = session.get('selected_branch_id')
    sales_data = get_summary_list_of_sales(year, month, branch_id=current_branch_id)

    columns = ['customer_tin', 'customer_name', 'customer_address', 'vatable_sales',
               'vat_exempt_sales', 'zero_rated_sales', 'vat_amount', 'gross_sales']
    headers = ['TIN', 'Customer Name', 'Address', 'Vatable Sales',
               'VAT-Exempt Sales', 'Zero-Rated Sales', 'Output VAT', 'Gross Sales']

    filename = f'BIR_Summary_Sales_{year}_{month:02d}.xlsx'
    title = f'Summary List of Sales - {get_month_name(month)} {year}'

    return export_to_excel(sales_data, columns, headers, filename, title)


@reports_bp.route('/reports/bir/purchases')
@login_required
@accountant_or_admin_required
def bir_purchases():
    """Summary List of Purchases (Annex B) - Monthly VAT Purchases"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    current_branch_id = session.get('selected_branch_id')
    purchases_data = get_summary_list_of_purchases(year, month, branch_id=current_branch_id)

    return render_template('reports/bir_purchases.html',
                         purchases_data=purchases_data,
                         year=year,
                         month=month,
                         month_name=get_month_name(month),
                         company=get_company_identity())


@reports_bp.route('/reports/bir/purchases/export/excel')
@login_required
@accountant_or_admin_required
def bir_purchases_export_excel():
    """Export Summary List of Purchases to Excel"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    current_branch_id = session.get('selected_branch_id')
    purchases_data = get_summary_list_of_purchases(year, month, branch_id=current_branch_id)

    columns = ['vendor_tin', 'vendor_name', 'vendor_address', 'vendor_invoice_number',
               'vatable_purchases', 'vat_exempt_purchases', 'zero_rated_purchases',
               'input_vat', 'gross_purchases']
    headers = ['TIN', 'Vendor Name', 'Address', 'Invoice #', 'Vatable Purchases',
               'VAT-Exempt Purchases', 'Zero-Rated Purchases', 'Input VAT', 'Gross Purchases']

    filename = f'BIR_Summary_Purchases_{year}_{month:02d}.xlsx'
    title = f'Summary List of Purchases - {get_month_name(month)} {year}'

    return export_to_excel(purchases_data, columns, headers, filename, title)


@reports_bp.route('/reports/bir/alphalist')
@login_required
@accountant_or_admin_required
def bir_alphalist():
    return redirect(url_for('dashboard.under_development', feature='BIR Alphalist'))
    year = request.args.get('year', datetime.now().year, type=int)
    quarter = request.args.get('quarter', ((datetime.now().month - 1) // 3) + 1, type=int)
    current_branch_id = session.get('selected_branch_id')
    payees_data = get_alphalist_of_payees(year, quarter, branch_id=current_branch_id)

    return render_template('reports/bir_alphalist.html',
                         payees_data=payees_data,
                         year=year,
                         quarter=quarter,
                         quarter_name=get_quarter_name(quarter),
                         quarter_months=get_quarter_months(quarter))


@reports_bp.route('/reports/bir/alphalist/export/excel')
@login_required
@accountant_or_admin_required
def bir_alphalist_export_excel():
    """Export Alphalist of Payees to Excel"""
    year = request.args.get('year', datetime.now().year, type=int)
    quarter = request.args.get('quarter', ((datetime.now().month - 1) // 3) + 1, type=int)
    current_branch_id = session.get('selected_branch_id')
    payees_data = get_alphalist_of_payees(year, quarter, branch_id=current_branch_id)

    columns = ['payee_tin', 'payee_name', 'payee_address', 'atc_code',
               'tax_rate', 'gross_income', 'tax_withheld', 'month_paid']
    headers = ['TIN', 'Payee Name', 'Address', 'ATC Code',
               'Tax Rate (%)', 'Gross Income', 'Tax Withheld', 'Month/s Paid']

    filename = f'BIR_Alphalist_Payees_{year}_Q{quarter}.xlsx'
    title = f'Alphalist of Payees - {get_quarter_name(quarter)} {year}'

    return export_to_excel(payees_data, columns, headers, filename, title)


# ============================================================================
# FINANCIAL STATEMENTS
# ============================================================================

def _tb_params():
    """Shared (as_of_date, branch_id) for Trial Balance & Balance Sheet routes.

    Defaults the reporting date to the end of the current month (PH time).
    """
    default = end_of_month(ph_now().date())
    as_of_str = request.args.get('as_of', default.isoformat())
    try:
        as_of_date = date.fromisoformat(as_of_str)
    except (ValueError, TypeError):
        as_of_date = default
    return as_of_date, session.get('selected_branch_id')


_TB_COLUMNS = ['code', 'name', 'debit_balance', 'credit_balance']
_TB_HEADERS = ['Account Code', 'Account Name', 'Debit', 'Credit']


@reports_bp.route('/reports/trial-balance')
@login_required
def trial_balance():
    as_of_date, branch_id = _tb_params()
    trial_balance_data = generate_trial_balance(as_of_date, branch_id=branch_id)
    return render_template('reports/trial_balance.html',
                           trial_balance=trial_balance_data,
                           as_of_date=as_of_date)


@reports_bp.route('/reports/trial-balance/export/excel')
@login_required
def trial_balance_export_excel():
    """Export Trial Balance to Excel"""
    as_of_date, branch_id = _tb_params()
    trial_balance_data = generate_trial_balance(as_of_date, branch_id=branch_id)
    filename = f'Trial_Balance_{as_of_date.isoformat()}.xlsx'
    title = f'Trial Balance - As of {as_of_date.strftime("%B %d, %Y")}'
    return export_to_excel(trial_balance_data['accounts'], _TB_COLUMNS, _TB_HEADERS, filename, title)


@reports_bp.route('/reports/trial-balance/print')
@login_required
def trial_balance_print():
    from app.settings import AppSettings
    from app.branches.models import Branch
    as_of_date, branch_id = _tb_params()
    trial_balance_data = generate_trial_balance(as_of_date, branch_id=branch_id)
    company = {
        'name': AppSettings.get_setting('company_name', ''),
        'address': AppSettings.get_setting('company_address', ''),
        'tin': AppSettings.get_setting('company_tin', ''),
    }
    branch = db.session.get(Branch, branch_id) if branch_id else None
    branch_name = branch.name if (branch and Branch.query.count() > 1) else None
    return render_template('reports/trial_balance_print.html',
                           trial_balance=trial_balance_data, as_of_date=as_of_date,
                           company=company, branch_name=branch_name)


def _is_params():
    """Shared (start_date, end_date, branch_id) parsing for Income Statement routes.

    Defaults to year-to-date (Jan 1 of the current year -> today).
    """
    today = date.today()

    def _parse(param, fallback):
        try:
            return date.fromisoformat(request.args.get(param, ''))
        except (ValueError, TypeError):
            return fallback

    return (_parse('start_date', date(today.year, 1, 1)),
            _parse('end_date', today),
            session.get('selected_branch_id'))


def _stmt_params():
    """Single reporting-date parsing for Income Statement & Cash Flow.

    Returns (as_of, mtd_start, ytd_start, branch_id). Reporting date defaults to
    the end of the current month (PH time); param name 'as_of'. A legacy
    'end_date' param is accepted as the reporting date for back-compat.
    """
    default = end_of_month(ph_now().date())
    raw = request.args.get('as_of') or request.args.get('end_date') or default.isoformat()
    try:
        as_of = date.fromisoformat(raw)
    except (ValueError, TypeError):
        as_of = default
    mtd_start = as_of.replace(day=1)
    ytd_start = date(as_of.year, 1, 1)
    return as_of, mtd_start, ytd_start, session.get('selected_branch_id')


@reports_bp.route('/reports/income-statement')
@login_required
def income_statement():
    as_of, mtd_start, ytd_start, branch_id = _stmt_params()
    mtd = generate_income_statement(mtd_start, as_of, branch_id=branch_id)
    ytd = generate_income_statement(ytd_start, as_of, branch_id=branch_id)
    data = merge_is_two_column(mtd, ytd)
    return render_template('reports/income_statement.html',
                           income_statement=data, as_of=as_of,
                           mtd_start=mtd_start, ytd_start=ytd_start)


@reports_bp.route('/reports/income-statement/export/excel')
@login_required
def income_statement_export_excel():
    """Export Income Statement to a formatted Excel workbook."""
    from app.settings import AppSettings
    from app.branches.models import Branch
    from app.reports.statement_export import build_income_statement_xlsx
    start_date, end_date, branch_id = _is_params()
    stmt = generate_income_statement(start_date, end_date, branch_id=branch_id)
    company = {
        'name': AppSettings.get_setting('company_name', ''),
        'address': AppSettings.get_setting('company_address', ''),
        'tin': AppSettings.get_setting('company_tin', ''),
    }
    branch = db.session.get(Branch, branch_id) if branch_id else None
    branch_name = branch.name if (branch and Branch.query.count() > 1) else None
    period_label = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    filename = f'Income_Statement_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx'
    return build_income_statement_xlsx(stmt, period_label, company, branch_name, filename)


@reports_bp.route('/reports/income-statement/print')
@login_required
def income_statement_print():
    from app.settings import AppSettings
    from app.branches.models import Branch
    from app.reports.statement_export import income_statement_lines
    start_date, end_date, branch_id = _is_params()
    income_stmt_data = generate_income_statement(start_date, end_date, branch_id=branch_id)
    company = {
        'name': AppSettings.get_setting('company_name', ''),
        'address': AppSettings.get_setting('company_address', ''),
        'tin': AppSettings.get_setting('company_tin', ''),
    }
    branch = db.session.get(Branch, branch_id) if branch_id else None
    branch_name = branch.name if (branch and Branch.query.count() > 1) else None
    return render_template('reports/income_statement_print.html',
                           lines=income_statement_lines(income_stmt_data),
                           start_date=start_date, end_date=end_date,
                           company=company, branch_name=branch_name)


def _bs_company_branch(branch_id):
    """(company dict, branch_name-or-None) for the Balance Sheet export/print headers."""
    from app.settings import AppSettings
    from app.branches.models import Branch
    company = {
        'name': AppSettings.get_setting('company_name', ''),
        'address': AppSettings.get_setting('company_address', ''),
        'tin': AppSettings.get_setting('company_tin', ''),
    }
    branch = db.session.get(Branch, branch_id) if branch_id else None
    branch_name = branch.name if (branch and Branch.query.count() > 1) else None
    return company, branch_name


@reports_bp.route('/reports/balance-sheet')
@login_required
def balance_sheet():
    as_of_date, branch_id = _tb_params()
    balance_sheet_data = generate_balance_sheet(as_of_date, branch_id=branch_id)
    return render_template('reports/balance_sheet.html',
                           balance_sheet=balance_sheet_data, as_of_date=as_of_date)


@reports_bp.route('/reports/balance-sheet/export/excel')
@login_required
def balance_sheet_export_excel():
    """Export Balance Sheet to a formatted Excel workbook."""
    from app.reports.statement_export import build_balance_sheet_xlsx
    as_of_date, branch_id = _tb_params()
    bs = generate_balance_sheet(as_of_date, branch_id=branch_id)
    company, branch_name = _bs_company_branch(branch_id)
    as_of_label = f'As of {as_of_date.strftime("%B %d, %Y")}'
    filename = f'Balance_Sheet_{as_of_date.isoformat()}.xlsx'
    return build_balance_sheet_xlsx(bs, as_of_label, company, branch_name, filename)


@reports_bp.route('/reports/balance-sheet/print')
@login_required
def balance_sheet_print():
    from app.reports.statement_export import balance_sheet_lines
    as_of_date, branch_id = _tb_params()
    bs = generate_balance_sheet(as_of_date, branch_id=branch_id)
    company, branch_name = _bs_company_branch(branch_id)
    return render_template('reports/balance_sheet_print.html',
                           lines=balance_sheet_lines(bs), as_of_date=as_of_date,
                           company=company, branch_name=branch_name)


@reports_bp.route('/reports/cash-flow')
@login_required
def cash_flow():
    start_date, end_date, branch_id = _is_params()
    cf = generate_cash_flow(start_date, end_date, branch_id=branch_id)
    return render_template('reports/cash_flow.html', cash_flow=cf,
                           start_date=start_date, end_date=end_date)


@reports_bp.route('/reports/cash-flow/export/excel')
@login_required
def cash_flow_export_excel():
    """Export the Statement of Cash Flows to a formatted Excel workbook."""
    from app.reports.statement_export import build_cash_flow_xlsx
    start_date, end_date, branch_id = _is_params()
    cf = generate_cash_flow(start_date, end_date, branch_id=branch_id)
    company, branch_name = _bs_company_branch(branch_id)
    period_label = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    filename = f'Cash_Flow_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx'
    return build_cash_flow_xlsx(cf, period_label, company, branch_name, filename)


@reports_bp.route('/reports/cash-flow/print')
@login_required
def cash_flow_print():
    from app.reports.statement_export import cash_flow_lines
    start_date, end_date, branch_id = _is_params()
    cf = generate_cash_flow(start_date, end_date, branch_id=branch_id)
    company, branch_name = _bs_company_branch(branch_id)
    return render_template('reports/cash_flow_print.html',
                           lines=cash_flow_lines(cf), start_date=start_date,
                           end_date=end_date, company=company, branch_name=branch_name)


def _descendant_leaf_ids(account):
    """IDs of all postable leaf accounts under `account` (recursive).

    Returns [] when `account` is itself a leaf (no children) — the caller then
    falls back to the single-account ledger path.
    """
    leaves = []
    stack = list(account.children)
    while stack:
        node = stack.pop()
        kids = list(node.children)
        if kids:
            stack.extend(kids)
        else:
            leaves.append(node.id)
    return leaves


def _aggregated_ledger_response(account, leaf_ids, start_date, end_date, branch_id):
    """Combined ledger of a parent group's descendant leaves, signed by the
    parent's normal_balance (the group's net movement). Each line is tagged
    with the child account it came from."""
    from app import db
    from app.journal_entries.models import JournalEntry, JournalEntryLine
    from sqlalchemy import func
    from decimal import Decimal

    sign = 1 if (account.normal_balance or 'debit').lower() == 'debit' else -1

    opening_raw = db.session.query(
        func.coalesce(func.sum(JournalEntryLine.debit_amount - JournalEntryLine.credit_amount), 0)
    ).join(JournalEntry).filter(
        JournalEntry.status == 'posted',
        JournalEntry.branch_id == branch_id,
        JournalEntry.entry_date < start_date,
        JournalEntryLine.account_id.in_(leaf_ids),
    ).scalar()
    running = Decimal(str(opening_raw or '0'))
    opening = float(running) * sign

    rows = db.session.query(JournalEntryLine, JournalEntry, Account).join(
        JournalEntry, JournalEntryLine.entry_id == JournalEntry.id
    ).join(Account, Account.id == JournalEntryLine.account_id).filter(
        JournalEntry.status == 'posted',
        JournalEntry.branch_id == branch_id,
        JournalEntry.entry_date >= start_date,
        JournalEntry.entry_date <= end_date,
        JournalEntryLine.account_id.in_(leaf_ids),
    ).order_by(
        JournalEntry.entry_date, JournalEntry.entry_number, JournalEntryLine.line_number
    ).all()

    lines = []
    for line, entry, acct in rows:
        running += (line.debit_amount - line.credit_amount)
        source = entry.display_number or entry.reference or entry.entry_number or ''
        lines.append({
            'date': entry.entry_date.isoformat() if hasattr(entry.entry_date, 'isoformat') else str(entry.entry_date),
            'source': source,
            'particulars': line.description or entry.description or '',
            'account': {'code': acct.code, 'name': acct.name},
            'debit': float(line.debit_amount),
            'credit': float(line.credit_amount),
            'balance': float(running) * sign,
        })

    return jsonify({
        'account': {'code': account.code, 'name': account.name},
        'aggregated': True,
        'lines': lines,
        'opening': opening,
        'closing': float(running) * sign,
    })


@reports_bp.route('/reports/account-ledger')
@login_required
def account_ledger_json():
    """JSON endpoint: per-account ledger for a date range (FS drill-down).

    GET /reports/account-ledger?account_id=<id>&start=<YYYY-MM-DD>&end=<YYYY-MM-DD>

    Returns:
        {
            'account': {'code': str, 'name': str},
            'lines': [{'date', 'source', 'particulars', 'debit', 'credit', 'balance'}],
            'opening': float,   # signed by normal_balance
            'closing': float,
        }
    Opening/closing are signed by the account's normal_balance:
        debit-normal  → debit − credit (positive = debit balance)
        credit-normal → credit − debit (positive = credit balance)
    """
    account_id = request.args.get('account_id', type=int)
    start_str = request.args.get('start', '')
    end_str = request.args.get('end', '')

    try:
        start_date = date.fromisoformat(start_str)
    except (ValueError, TypeError):
        start_date = date(date.today().year, 1, 1)

    try:
        end_date = date.fromisoformat(end_str)
    except (ValueError, TypeError):
        end_date = date.today()

    if not account_id:
        return jsonify({'error': 'account_id is required'}), 400

    account = db.session.get(Account, account_id)
    if account is None:
        return jsonify({'error': 'Account not found'}), 404

    branch_id = session.get('selected_branch_id')

    # Parent (group) accounts have no postings of their own — aggregate the
    # ledgers of all postable descendant leaves into one combined ledger.
    leaf_ids = _descendant_leaf_ids(account)
    if leaf_ids:
        return _aggregated_ledger_response(account, leaf_ids, start_date, end_date, branch_id)

    # Reuse generate_general_ledger for the single-account ledger.
    # When branch_id is None the helper filters branch_id == None (no-branch JEs).
    # Mirror the GL report's behaviour: pass branch_id from session unchanged.
    from app.reports.financial import generate_general_ledger
    ledger = generate_general_ledger(start_date, end_date, branch_id, account_id=account_id)

    # generate_general_ledger stores opening/running in raw debit-positive form.
    # Apply normal_balance sign convention so callers get a meaningful balance sign.
    normal_balance = (account.normal_balance or 'debit').lower()
    sign = 1 if normal_balance == 'debit' else -1

    acct_data = ledger['accounts'][0] if ledger['accounts'] else None
    if acct_data is None:
        # No activity at all for this account in this period.
        return jsonify({
            'account': {'code': account.code, 'name': account.name},
            'aggregated': False,
            'lines': [],
            'opening': 0.0,
            'closing': 0.0,
        })

    opening = float(acct_data['opening_balance']) * sign

    lines = []
    for gl_line in acct_data['lines']:
        # Derive source label: use display_number (JV/reversal) or reference (auto-posted docs)
        source = gl_line.get('display_number') or gl_line.get('reference') or gl_line.get('entry_number', '')
        lines.append({
            'date': gl_line['entry_date'].isoformat() if hasattr(gl_line['entry_date'], 'isoformat') else str(gl_line['entry_date']),
            'source': source,
            'particulars': gl_line.get('description', ''),
            'debit': gl_line['debit'],
            'credit': gl_line['credit'],
            'balance': float(gl_line['running_balance']) * sign,
        })

    closing = float(acct_data['closing_balance']) * sign

    return jsonify({
        'account': {'code': account.code, 'name': account.name},
        'aggregated': False,
        'lines': lines,
        'opening': opening,
        'closing': closing,
    })


@reports_bp.route('/reports/ap-aging/export/excel')
@login_required
def ap_aging_export_excel():
    as_of_str = request.args.get('as_of', date.today().isoformat())
    try:
        as_of_date = date.fromisoformat(as_of_str)
    except ValueError:
        as_of_date = date.today()
    current_branch_id = session.get('selected_branch_id')
    vendors_list, grand_totals = _build_ap_aging_data(as_of_date, current_branch_id)
    rows = [
        {
            'name': v['name'],
            'current': v['current'],
            '1-30': v['1-30'],
            '31-60': v['31-60'],
            '61-90': v['61-90'],
            '90+': v['90+'],
            'total': v['total'],
        }
        for v in vendors_list
    ]
    rows.append({
        'name': 'GRAND TOTAL',
        'current': grand_totals['current'],
        '1-30': grand_totals['1-30'],
        '31-60': grand_totals['31-60'],
        '61-90': grand_totals['61-90'],
        '90+': grand_totals['90+'],
        'total': grand_totals['total'],
    })
    columns = ['name', 'current', '1-30', '31-60', '61-90', '90+', 'total']
    headers = ['Vendor', 'Current', '1-30', '31-60', '61-90', '90+', 'Total']
    return export_to_excel(rows, columns, headers,
                           filename=f'ap_aging_{as_of_date.isoformat()}.xlsx',
                           title=f'AP Aging as of {as_of_date}')


@reports_bp.route('/reports/ap-aging/export/csv')
@login_required
def ap_aging_export_csv():
    as_of_str = request.args.get('as_of', date.today().isoformat())
    try:
        as_of_date = date.fromisoformat(as_of_str)
    except ValueError:
        as_of_date = date.today()
    current_branch_id = session.get('selected_branch_id')
    vendors_list, grand_totals = _build_ap_aging_data(as_of_date, current_branch_id)
    rows = [
        {
            'name': v['name'],
            'current': v['current'],
            '1-30': v['1-30'],
            '31-60': v['31-60'],
            '61-90': v['61-90'],
            '90+': v['90+'],
            'total': v['total'],
        }
        for v in vendors_list
    ]
    rows.append({
        'name': 'GRAND TOTAL',
        'current': grand_totals['current'],
        '1-30': grand_totals['1-30'],
        '31-60': grand_totals['31-60'],
        '61-90': grand_totals['61-90'],
        '90+': grand_totals['90+'],
        'total': grand_totals['total'],
    })
    columns = ['name', 'current', '1-30', '31-60', '61-90', '90+', 'total']
    headers = ['Vendor', 'Current', '1-30', '31-60', '61-90', '90+', 'Total']
    return export_to_csv(rows, columns, headers,
                         filename=f'ap_aging_{as_of_date.isoformat()}.csv')


@reports_bp.route('/reports/ar-aging/export/excel')
@login_required
def ar_aging_export_excel():
    as_of_str = request.args.get('as_of', date.today().isoformat())
    try:
        as_of_date = date.fromisoformat(as_of_str)
    except ValueError:
        as_of_date = date.today()
    current_branch_id = session.get('selected_branch_id')
    customers_list, grand_totals = _build_ar_aging_data(as_of_date, current_branch_id)
    rows = [
        {
            'name': c['name'],
            'current': c['current'],
            '1-30': c['1-30'],
            '31-60': c['31-60'],
            '61-90': c['61-90'],
            '90+': c['90+'],
            'total': c['total'],
        }
        for c in customers_list
    ]
    rows.append({
        'name': 'GRAND TOTAL',
        'current': grand_totals['current'],
        '1-30': grand_totals['1-30'],
        '31-60': grand_totals['31-60'],
        '61-90': grand_totals['61-90'],
        '90+': grand_totals['90+'],
        'total': grand_totals['total'],
    })
    columns = ['name', 'current', '1-30', '31-60', '61-90', '90+', 'total']
    headers = ['Customer', 'Current', '1-30', '31-60', '61-90', '90+', 'Total']
    return export_to_excel(rows, columns, headers,
                           filename=f'ar_aging_{as_of_date.isoformat()}.xlsx',
                           title=f'AR Aging as of {as_of_date}')


@reports_bp.route('/reports/ar-aging/export/csv')
@login_required
def ar_aging_export_csv():
    as_of_str = request.args.get('as_of', date.today().isoformat())
    try:
        as_of_date = date.fromisoformat(as_of_str)
    except ValueError:
        as_of_date = date.today()
    current_branch_id = session.get('selected_branch_id')
    customers_list, grand_totals = _build_ar_aging_data(as_of_date, current_branch_id)
    rows = [
        {
            'name': c['name'],
            'current': c['current'],
            '1-30': c['1-30'],
            '31-60': c['31-60'],
            '61-90': c['61-90'],
            '90+': c['90+'],
            'total': c['total'],
        }
        for c in customers_list
    ]
    rows.append({
        'name': 'GRAND TOTAL',
        'current': grand_totals['current'],
        '1-30': grand_totals['1-30'],
        '31-60': grand_totals['31-60'],
        '61-90': grand_totals['61-90'],
        '90+': grand_totals['90+'],
        'total': grand_totals['total'],
    })
    columns = ['name', 'current', '1-30', '31-60', '61-90', '90+', 'total']
    headers = ['Customer', 'Current', '1-30', '31-60', '61-90', '90+', 'Total']
    return export_to_csv(rows, columns, headers,
                         filename=f'ar_aging_{as_of_date.isoformat()}.csv')


# ============================================================================
# BOOKS OF ACCOUNTS — General Journal
# ============================================================================

def _general_journal_entries(branch_id, period):
    """Voucher-type journal entries for the given branch and period, ordered by date+id."""
    return JournalEntry.query.filter(
        JournalEntry.entry_type.in_(VOUCHER_ENTRY_TYPES),
        JournalEntry.branch_id == branch_id,
        JournalEntry.entry_date >= period['date_from'],
        JournalEntry.entry_date <= period['date_to'],
    ).order_by(JournalEntry.entry_date, JournalEntry.id).all()


@reports_bp.route('/reports/general-journal')
@login_required
def general_journal():
    """General Journal screen view — voucher entries only."""
    branch_id = session.get('selected_branch_id')
    if not branch_id:
        flash('Please select a branch to view the General Journal.', 'warning')
        return redirect(url_for('users.select_branch', next=request.url))
    period = resolve_period(request.args, ph_now().date())
    gj = build_general_journal(_general_journal_entries(branch_id, period))
    return render_template('reports/general_journal.html', gj=gj, period=period)


@reports_bp.route('/reports/general-journal/print')
@login_required
def general_journal_print():
    """General Journal printable view."""
    branch_id = session.get('selected_branch_id')
    if not branch_id:
        flash('Please select a branch.', 'warning')
        return redirect(url_for('users.select_branch', next=request.url))
    period = resolve_period(request.args, ph_now().date())
    gj = build_general_journal(_general_journal_entries(branch_id, period))
    return render_template('reports/general_journal_print.html', gj=gj, period=period,
                           company=get_company_identity(), printed_at=ph_now())


@reports_bp.route('/reports/general-journal/export')
@login_required
def general_journal_export():
    """General Journal Excel export."""
    from flask import send_file
    from app.branches.models import Branch
    branch_id = session.get('selected_branch_id')
    if not branch_id:
        flash('Please select a branch.', 'warning')
        return redirect(url_for('users.select_branch', next=request.url))
    period = resolve_period(request.args, ph_now().date())
    gj = build_general_journal(_general_journal_entries(branch_id, period))
    branch = db.session.get(Branch, branch_id)
    branch_name = branch.name if branch else None
    bio = build_general_journal_xlsx(gj, period['label'], get_company_identity(),
                                     branch_name, 'General_Journal.xlsx')
    return send_file(bio, as_attachment=True, download_name='General_Journal.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================================
# BOOKS OF ACCOUNTS — Hub
# ============================================================================

@reports_bp.route('/reports/books-of-accounts')
@login_required
@accountant_or_admin_required
def books_of_accounts():
    """Hub page listing all six BIR books of accounts with per-book View links."""
    year = ph_now().year
    date_from = request.args.get('date_from', f'{year}-01-01')
    date_to = request.args.get('date_to', f'{year}-12-31')
    return render_template('reports/books_of_accounts.html',
                           books=BOOKS, date_from=date_from, date_to=date_to)


@reports_bp.route('/reports/books-of-accounts/print-all')
@login_required
@accountant_or_admin_required
def books_print_all():
    """Render all six BIR books stacked in a single print-ready page."""
    from app.reports.books_data import collect_books
    branch_id = session.get('selected_branch_id')
    if not branch_id:
        flash('Please select a branch.', 'warning')
        return redirect(url_for('users.select_branch', next=request.url))
    bundle = collect_books(branch_id, request.args)
    # Build ordered list so the template can iterate with loop.first
    books = [{'title': m['title'], **bundle['books'][m['key']]} for m in BOOKS]
    return render_template(
        'reports/books_print_all.html',
        company=get_company_identity(),
        period=bundle['period'],
        books=books,
        printed_at=ph_now(),
    )


@reports_bp.route('/reports/books-of-accounts/export-all')
@login_required
@accountant_or_admin_required
def books_export_all():
    """One openpyxl workbook with one sheet per BIR book."""
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from app.reports.books_data import collect_books
    from app.utils.bir_books import write_bir_book_header

    branch_id = session.get('selected_branch_id')
    if not branch_id:
        flash('Please select a branch.', 'warning')
        return redirect(url_for('users.select_branch', next=request.url))

    bundle = collect_books(branch_id, request.args)
    company = get_company_identity()
    period_label = bundle['period']['label']
    NUM = '#,##0.00;(#,##0.00)'

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for meta in BOOKS:
        ws = wb.create_sheet(meta['sheet'])
        book = bundle['books'][meta['key']]
        write_bir_book_header(ws, company, meta['title'].upper(), period_label)

        if book['kind'] == 'gj':
            gj = book['data']
            ws.append(['Date', 'Particulars', 'Ref', 'Debit', 'Credit'])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            _write_gj_rows(ws, gj)
            for col_letter in ('D', 'E'):
                for cell in ws[col_letter]:
                    cell.number_format = NUM

        elif book['kind'] == 'columnar':
            m = book['data']
            header = ['Date', 'Ref', 'Particulars'] + [col['name'] for col in m['columns']]
            ws.append(header)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for r in m['rows']:
                if r.get('is_voided') and r.get('entry') is None:
                    # Voided row — emit a [VOIDED] marker row mirroring the print
                    # template's col-row-voided rendering: date from the source
                    # document, "[VOIDED]" in the Ref column, doc number in
                    # Particulars, blank/zero amounts.  Sales Journal rows carry
                    # key 'invoice'; Purchase Journal rows carry key 'ap'.
                    inv = r.get('invoice')
                    ap = r.get('ap')
                    if inv:
                        date_str = inv.invoice_date.strftime('%m/%d/%Y') if inv.invoice_date else ''
                        particulars = inv.invoice_number or ''
                    elif ap:
                        date_str = ap.ap_date.strftime('%m/%d/%Y') if ap.ap_date else ''
                        particulars = ap.ap_number or ''
                    else:
                        date_str = ''
                        particulars = ''
                    ws.append([date_str, '[VOIDED]', particulars] + [None] * len(m['columns']))
                    continue
                if r.get('is_cancelled'):
                    # CR/CD cancelled rows: JE exists; reference is the CRV/CDV number.
                    e = r['entry']
                    date_str = e.entry_date.strftime('%m/%d/%Y') if e.entry_date else ''
                    ws.append([date_str, '[VOID]', e.reference or ''] + [None] * len(m['columns']))
                    continue
                e = r['entry']
                ws.append(
                    [e.entry_date.strftime('%m/%d/%Y'), e.display_number, e.description or ''] +
                    [float(r['cells'].get(col['account_id'], 0) or 0) for col in m['columns']]
                )
            totals_row = ['', '', 'TOTAL'] + [
                float(m['totals'].get(col['account_id'], 0) or 0)
                for col in m['columns']
            ]
            ws.append(totals_row)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            # Apply number format to amount columns (D onward)
            for col_idx in range(4, 4 + len(m['columns'])):
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter]:
                    cell.number_format = NUM

        else:  # 'gl' — pointer row; full detail is available via the standalone GL export
            ws.append(['See the General Ledger report for full per-account detail.'])
            ws.cell(ws.max_row, 1).font = Font(italic=True)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name='Books_of_Accounts.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
