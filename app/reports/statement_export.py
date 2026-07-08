"""Line-flattening + professional .xlsx builders for the financial statements.

`income_statement_lines` is the single source of the printed/exported P&L layout
(shared by the print template and the Excel builder) so they stay identical.
"""
import io

from flask import make_response, has_request_context
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

_XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
_NUM_FMT = '#,##0.00;(#,##0.00)'   # accounting style: parentheses for negatives


def income_statement_lines(stmt):
    """Flatten the two-column P&L into render-ready lines (shared by print + Excel).

    Consumes the merged two-column shape from merge_is_two_column():
      stmt['sections'] = [{'key','label','sign','mtd_total','ytd_total','lines':[...],
                           optional 'subtotal_label','mtd_subtotal','ytd_subtotal'}, ...]
    Each group/child line carries 'mtd_amount'/'ytd_amount'.

    Each returned line: {'kind', 'label', 'mtd' (or None), 'ytd' (or None),
    'indent', 'rule'}. rule ∈ {None, 'bottom', 'top_bottom', 'double_bottom'}.
    """
    lines = []
    for sec in stmt['sections']:
        header = sec['label']
        groups = sec['lines']   # {code, name, account_id, mtd_amount, ytd_amount, children:[...]}
        if not groups:
            # Empty section: one ruled total line
            lines.append({'kind': 'total', 'label': header,
                          'mtd': sec['mtd_total'], 'ytd': sec['ytd_total'],
                          'indent': False, 'rule': 'bottom'})
        elif len(groups) == 1 and not groups[0]['children']:
            # Single leaf account — header + account line (no separate total)
            lines.append({'kind': 'header', 'label': header, 'mtd': None, 'ytd': None,
                          'indent': False, 'rule': None})
            g = groups[0]
            lines.append({'kind': 'account',
                          'label': f"{g['code']}  {g['name']}" if g['code'] else g['name'],
                          'mtd': g['mtd_amount'], 'ytd': g['ytd_amount'],
                          'indent': True, 'rule': 'bottom'})
        else:
            # Multiple groups or groups with children
            lines.append({'kind': 'header', 'label': header, 'mtd': None, 'ytd': None,
                          'indent': False, 'rule': None})
            for g in groups:
                if g['children']:
                    # Parent group: show group header then indented children
                    lines.append({'kind': 'account',
                                  'label': f"{g['code']}  {g['name']}" if g['code'] else g['name'],
                                  'mtd': None, 'ytd': None, 'indent': True, 'rule': None})
                    for c in g['children']:
                        lines.append({'kind': 'account',
                                      'label': f"{c['code']}  {c['name']}" if c['code'] else c['name'],
                                      'mtd': c['mtd_amount'], 'ytd': c['ytd_amount'],
                                      'indent': True, 'rule': None})
                else:
                    # Leaf group (no children)
                    lines.append({'kind': 'account',
                                  'label': f"{g['code']}  {g['name']}" if g['code'] else g['name'],
                                  'mtd': g['mtd_amount'], 'ytd': g['ytd_amount'],
                                  'indent': True, 'rule': None})
            lines.append({'kind': 'total', 'label': 'Total ' + header,
                          'mtd': sec['mtd_total'], 'ytd': sec['ytd_total'],
                          'indent': False, 'rule': 'top_bottom'})

        # Subtotal row after the section (e.g. Gross Profit, Net Income)
        if sec.get('subtotal_label'):
            is_net_income = sec['subtotal_label'] == 'Net Income'
            lines.append({'kind': 'net' if is_net_income else 'subtotal',
                          'label': sec['subtotal_label'],
                          'mtd': sec['mtd_subtotal'], 'ytd': sec['ytd_subtotal'],
                          'indent': False,
                          'rule': 'double_bottom' if is_net_income else 'bottom'})

    return lines


def _wb_bytes(wb):
    """Serialize workbook to raw bytes."""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _xlsx_response(wb, filename):
    """Return an HTTP response with the workbook bytes (requires request context)."""
    data = _wb_bytes(wb)
    resp = make_response(data)
    resp.headers['Content-Type'] = _XLSX_MIME
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _xlsx_response_or_bytes(wb, filename):
    """Return HTTP response in a request context, raw bytes otherwise (e.g. tests)."""
    if has_request_context():
        return _xlsx_response(wb, filename)
    return _wb_bytes(wb)


def build_income_statement_xlsx(stmt, as_of_label, company, branch_name, filename):
    """Render the two-column Income Statement as a formatted workbook.

    Consumes the merged two-column shape from merge_is_two_column(); writes a
    Particulars column plus a current-month and a year-to-date amount column.
    Returns raw bytes (BytesIO content) when called without a Flask request context,
    or an HTTP response when called from a view.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Income Statement'

    right = Alignment(horizontal='right')
    thin, double_s = Side(style='thin'), Side(style='double')
    rules = {
        'bottom': Border(bottom=thin),
        'top_bottom': Border(top=thin, bottom=thin),
        'double_bottom': Border(bottom=double_s),
    }

    def put(particulars='', mtd=None, ytd=None):
        ws.append([particulars, mtd, ytd])
        return ws.max_row

    # ── Header ──────────────────────────────────────────────────────────────
    if company.get('name'):
        r = put(company['name']); ws.cell(r, 1).font = Font(bold=True, size=14)
    meta = []
    if company.get('tin'):
        meta.append('TIN: ' + company['tin'])
    if company.get('address'):
        meta.append(company['address'])
    if meta:
        put(' · '.join(meta))
    if branch_name:
        put('Branch: ' + branch_name)
    r = put('Income Statement (Profit & Loss)'); ws.cell(r, 1).font = Font(bold=True, size=13)
    put(as_of_label)
    put()

    as_of = stmt['as_of']
    r = put('Particulars', as_of.strftime('%b %Y'), f'YTD {as_of.year}')
    for cell in ws[r]:
        cell.font = Font(bold=True)
        cell.border = Border(bottom=thin)
    ws.cell(r, 2).alignment = right
    ws.cell(r, 3).alignment = right

    # ── Body (reuse the shared flattener) ─────────────────────────────────────
    def style(r, bold=False, border=None):
        font = Font(bold=True) if bold else None
        for col in (1, 2, 3):
            c = ws.cell(r, col)
            if font:
                c.font = font
            if border:
                c.border = border
            if col in (2, 3):
                c.number_format = _NUM_FMT
                c.alignment = right

    for ln in income_statement_lines(stmt):
        indent = '        ' if ln['indent'] else ''
        r = put(indent + ln['label'], ln['mtd'], ln['ytd'])
        style(r, bold=ln['kind'] in ('header', 'total', 'subtotal', 'net'),
              border=rules.get(ln['rule']))

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.sheet_view.showGridLines = False

    return _xlsx_response_or_bytes(wb, filename)


_BS_GRAND = {'assets': 'TOTAL ASSETS', 'liabilities': 'TOTAL LIABILITIES', 'equity': 'TOTAL EQUITY'}


def balance_sheet_lines(bs):
    """Flatten the classified balance sheet into render-ready lines (print + Excel).

    Consumes the type-driven shape from generate_balance_sheet():
      bs['sections'] = [{'key','label','total',
                          'divisions':[{'label','total','lines':[...]}]}, ...]

    Each returned line: {'kind', 'label', 'amount' (or None), 'indent', 'rule'}.
    indent ∈ {0, 1, 2} — matches the CSS indent classes in the print template.
    rule ∈ {None, 'bottom', 'top_bottom', 'double_bottom'}.
    """
    lines = []
    for sec in bs['sections']:
        lines.append({'kind': 'section', 'label': sec['label'], 'amount': None,
                      'indent': 0, 'rule': None})
        nonempty_divs = [d for d in sec['divisions'] if d['total'] != 0 or d['lines']]
        single_div = len(nonempty_divs) == 1
        for div in sec['divisions']:
            # Skip empty divisions
            if div['total'] == 0 and not div['lines']:
                continue
            if not single_div:
                lines.append({'kind': 'group', 'label': div['label'], 'amount': None,
                              'indent': 1, 'rule': None})
            for g in div['lines']:
                indent_account = 1 if single_div else 2
                if g['children']:
                    # Group header (no amount)
                    nm = f"{g['code']}  {g['name']}" if g['code'] else g['name']
                    lines.append({'kind': 'account', 'label': nm, 'amount': None,
                                  'indent': indent_account, 'rule': None})
                    for c in g['children']:
                        cnm = f"{c['code']}  {c['name']}" if c['code'] else c['name']
                        lines.append({'kind': 'account', 'label': cnm, 'amount': c['amount'],
                                      'indent': indent_account, 'rule': None})
                else:
                    nm = f"{g['code']}  {g['name']}" if g['code'] else g['name']
                    lines.append({'kind': 'account', 'label': nm, 'amount': g['total'],
                                  'indent': indent_account, 'rule': None})
            if not single_div:
                lines.append({'kind': 'group_total', 'label': 'Total ' + div['label'],
                              'amount': div['total'], 'indent': 1, 'rule': 'top_bottom'})
        rule = 'double_bottom' if sec['key'] == 'assets' else 'bottom'
        lines.append({'kind': 'section_total', 'label': _BS_GRAND[sec['key']],
                      'amount': sec['total'], 'indent': 0, 'rule': rule})
    lines.append({'kind': 'grand_total', 'label': 'TOTAL LIABILITIES AND EQUITY',
                  'amount': bs['total_liabilities_equity'], 'indent': 0, 'rule': 'double_bottom'})
    return lines


def build_balance_sheet_xlsx(bs, as_of_label, company, branch_name, filename):
    """Classified Balance Sheet as a formatted workbook with live SUM formulas.

    Consumes the type-driven shape from generate_balance_sheet().
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Balance Sheet'

    right = Alignment(horizontal='right')
    thin, double_s = Side(style='thin'), Side(style='double')
    rules = {
        'bottom': Border(bottom=thin),
        'top_bottom': Border(top=thin, bottom=thin),
        'double_bottom': Border(bottom=double_s),
    }

    def put(label='', amount=None):
        ws.append([label, amount])
        return ws.max_row

    def style(r, bold=False, size=None, border=None):
        font = Font(bold=bold, size=size) if size else (Font(bold=True) if bold else None)
        lc = ws.cell(r, 1)
        if font:
            lc.font = font
        if border:
            lc.border = border
        ac = ws.cell(r, 2)
        ac.number_format = _NUM_FMT
        ac.alignment = right
        if font:
            ac.font = font
        if border:
            ac.border = border

    # ── Header ──
    if company.get('name'):
        r = put(company['name']); ws.cell(r, 1).font = Font(bold=True, size=14)
    meta = []
    if company.get('tin'):
        meta.append('TIN: ' + company['tin'])
    if company.get('address'):
        meta.append(company['address'])
    if meta:
        put(' · '.join(meta))
    if branch_name:
        put('Branch: ' + branch_name)
    r = put('Balance Sheet'); ws.cell(r, 1).font = Font(bold=True, size=13)
    put(as_of_label)
    put()
    r = put('Particulars', 'Amount')
    for cell in ws[r]:
        cell.font = Font(bold=True)
        cell.border = Border(bottom=thin)
    ws.cell(r, 2).alignment = right

    # ── Body ─────────────────────────────────────────────────────────────────
    section_total_row = {}
    for sec in bs['sections']:
        r = put(sec['label']); ws.cell(r, 1).font = Font(bold=True)   # section header

        nonempty_divs = [d for d in sec['divisions'] if d['total'] != 0 or d['lines']]
        single_div = len(nonempty_divs) == 1
        div_total_rows = []

        for div in sec['divisions']:
            # Skip empty divisions for clean output
            if div['total'] == 0 and not div['lines']:
                continue
            if not single_div:
                r = put('    ' + div['label']); ws.cell(r, 1).font = Font(bold=True)

            # Emit account rows for each group in this division
            acct_rows = []
            for g in div['lines']:
                indent = '        ' if not single_div else '    '
                if g['children']:
                    # Group header (no amount)
                    nm = f"{g['code']}  {g['name']}" if g['code'] else g['name']
                    r = put(indent + nm); ws.cell(r, 1).font = Font(bold=True)
                    for c in g['children']:
                        cnm = f"{c['code']}  {c['name']}" if c['code'] else c['name']
                        extra = '    '
                        r = put(indent + extra + cnm, c['amount']); style(r)
                        acct_rows.append(r)
                else:
                    nm = f"{g['code']}  {g['name']}" if g['code'] else g['name']
                    r = put(indent + nm, g['total']); style(r)
                    acct_rows.append(r)

            if not single_div:
                r = put('    Total ' + div['label'])
                ws.cell(r, 2).value = f'=SUM(B{acct_rows[0]}:B{acct_rows[-1]})' if acct_rows else 0
                style(r, bold=True, border=rules['top_bottom'])
                div_total_rows.append(r)
            else:
                # Single division: section total sums the account rows directly
                div_total_rows = acct_rows

        st = put(_BS_GRAND[sec['key']])
        if div_total_rows:
            if single_div:
                ws.cell(st, 2).value = f'=SUM(B{div_total_rows[0]}:B{div_total_rows[-1]})' if div_total_rows else 0
            else:
                ws.cell(st, 2).value = '=' + '+'.join(f'B{x}' for x in div_total_rows)
        else:
            ws.cell(st, 2).value = 0
        style(st, bold=True, border=rules['double_bottom' if sec['key'] == 'assets' else 'bottom'])
        section_total_row[sec['key']] = st

    gt = put('TOTAL LIABILITIES AND EQUITY')
    liab_row = section_total_row.get('liabilities')
    eq_row = section_total_row.get('equity')
    if liab_row and eq_row:
        ws.cell(gt, 2).value = f"=B{liab_row}+B{eq_row}"
    else:
        ws.cell(gt, 2).value = bs['total_liabilities_equity']
    style(gt, bold=True, size=12, border=rules['double_bottom'])

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 22
    ws.sheet_view.showGridLines = False

    return _xlsx_response_or_bytes(wb, filename)


def _cf_invfin_net_cash_lines(cf):
    """Investing + Financing sections + NET INCREASE + begin/end cash.

    Shared by both methods so the tail is emitted identically.
    """
    lines = []
    for key, label, short in (('investing', 'CASH FLOWS FROM INVESTING ACTIVITIES', 'investing'),
                              ('financing', 'CASH FLOWS FROM FINANCING ACTIVITIES', 'financing')):
        sec = cf[key]
        lines.append({'kind': 'header', 'label': label, 'amount': None, 'indent': False, 'rule': None})
        for ln in sec['lines']:
            lines.append({'kind': 'account', 'label': ln['name'], 'amount': ln['amount'],
                          'indent': True, 'rule': None})
        lines.append({'kind': 'subtotal',
                      'label': 'Net cash provided by/(used in) %s activities' % short,
                      'amount': sec['total'], 'indent': False, 'rule': 'top_bottom'})
    lines.append({'kind': 'net', 'label': 'NET INCREASE/(DECREASE) IN CASH',
                  'amount': cf['net_change'], 'indent': False, 'rule': 'double_bottom'})
    lines.append({'kind': 'total', 'label': 'Cash at beginning of period',
                  'amount': cf['cash_begin'], 'indent': False, 'rule': None})
    lines.append({'kind': 'total', 'label': 'Cash at end of period',
                  'amount': cf['cash_end'], 'indent': False, 'rule': 'double_bottom'})
    return lines


def _reconciliation_lines(rec):
    """The PAS 7 net-income -> operating-cash reconciliation note."""
    lines = [{'kind': 'subheader',
              'label': 'Reconciliation of net income to net cash from operating activities',
              'amount': None, 'indent': False, 'rule': None},
             {'kind': 'account', 'label': 'Net Income (period)',
              'amount': rec['net_income'], 'indent': True, 'rule': None}]
    if rec['depreciation']:
        lines.append({'kind': 'account', 'label': 'Add: Depreciation',
                      'amount': rec['depreciation'], 'indent': True, 'rule': None})
    for w in rec['working_capital']:
        lines.append({'kind': 'account', 'label': w['name'], 'amount': w['amount'],
                      'indent': True, 'rule': None})
    lines.append({'kind': 'subtotal', 'label': 'Net cash from operating activities',
                  'amount': rec['total'], 'indent': False, 'rule': 'top_bottom'})
    return lines


def cash_flow_lines(cf):
    """Flatten the cash flow statement into render-ready lines (print + Excel)."""
    if cf.get('method') == 'direct':
        lines = [{'kind': 'header', 'label': 'CASH FLOWS FROM OPERATING ACTIVITIES',
                  'amount': None, 'indent': False, 'rule': None}]
        for ln in cf['operating']['lines']:
            lines.append({'kind': 'account', 'label': ln['name'], 'amount': ln['amount'],
                          'indent': True, 'rule': None})
        lines.append({'kind': 'subtotal',
                      'label': 'Net cash provided by/(used in) operating activities',
                      'amount': cf['operating']['total'], 'indent': False, 'rule': 'top_bottom'})
        lines += _cf_invfin_net_cash_lines(cf)
        if cf.get('noncash'):
            lines.append({'kind': 'subheader',
                          'label': 'Non-cash investing and financing transactions',
                          'amount': None, 'indent': False, 'rule': None})
            for n in cf['noncash']:
                lines.append({'kind': 'account', 'label': n['description'],
                              'amount': n['amount'], 'indent': True, 'rule': None})
        lines += _reconciliation_lines(cf['reconciliation'])
        return lines

    # Indirect (unchanged output)
    lines = []
    op = cf['operating']
    lines.append({'kind': 'header', 'label': 'CASH FLOWS FROM OPERATING ACTIVITIES',
                  'amount': None, 'indent': False, 'rule': None})
    lines.append({'kind': 'account', 'label': 'Net Income (period)',
                  'amount': op['net_income'], 'indent': True, 'rule': None})
    if op['depreciation']:
        lines.append({'kind': 'account', 'label': 'Add: Depreciation',
                      'amount': op['depreciation'], 'indent': True, 'rule': None})
    if op['working_capital']:
        lines.append({'kind': 'subheader', 'label': 'Changes in operating assets and liabilities:',
                      'amount': None, 'indent': True, 'rule': None})
        for w in op['working_capital']:
            lines.append({'kind': 'account', 'label': w['name'], 'amount': w['amount'],
                          'indent': True, 'rule': None})
    lines.append({'kind': 'subtotal', 'label': 'Net cash provided by/(used in) operating activities',
                  'amount': op['total'], 'indent': False, 'rule': 'top_bottom'})
    lines += _cf_invfin_net_cash_lines(cf)
    return lines


def build_cash_flow_xlsx(cf, period_label, company, branch_name, filename):
    """Statement of Cash Flows as a formatted workbook with live formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cash Flow'

    right = Alignment(horizontal='right')
    thin, double_s = Side(style='thin'), Side(style='double')
    rules = {
        'bottom': Border(bottom=thin),
        'top_bottom': Border(top=thin, bottom=thin),
        'double_bottom': Border(bottom=double_s),
    }

    def put(label='', amount=None):
        ws.append([label, amount])
        return ws.max_row

    def style(r, bold=False, size=None, border=None):
        font = Font(bold=bold, size=size) if size else (Font(bold=True) if bold else None)
        lc = ws.cell(r, 1)
        if font:
            lc.font = font
        if border:
            lc.border = border
        ac = ws.cell(r, 2)
        ac.number_format = _NUM_FMT
        ac.alignment = right
        if font:
            ac.font = font
        if border:
            ac.border = border

    is_direct = cf.get('method') == 'direct'

    # Header
    if company.get('name'):
        r = put(company['name']); ws.cell(r, 1).font = Font(bold=True, size=14)
    meta = []
    if company.get('tin'):
        meta.append('TIN: ' + company['tin'])
    if company.get('address'):
        meta.append(company['address'])
    if meta:
        put(' · '.join(meta))
    if branch_name:
        put('Branch: ' + branch_name)
    r = put('Statement of Cash Flows'); ws.cell(r, 1).font = Font(bold=True, size=13)
    put('Direct Method' if is_direct else 'Indirect Method')
    put(period_label)
    put()
    r = put('Particulars', 'Amount')
    for cell in ws[r]:
        cell.font = Font(bold=True)
        cell.border = Border(bottom=thin)
    ws.cell(r, 2).alignment = right

    # Operating section (live SUM over its detail rows)
    r = put('CASH FLOWS FROM OPERATING ACTIVITIES'); ws.cell(r, 1).font = Font(bold=True)
    if is_direct:
        first = last = None
        for ln in cf['operating']['lines']:
            r = put('        ' + ln['name'], ln['amount']); style(r)
            first = first or r
            last = r
    else:
        op = cf['operating']
        r = put('        Net Income (period)', op['net_income']); style(r)
        first = last = r
        if op['depreciation']:
            r = put('        Add: Depreciation', op['depreciation']); style(r); last = r
        if op['working_capital']:
            r = put('        Changes in operating assets and liabilities:')
            ws.cell(r, 1).font = Font(italic=True)
            for w in op['working_capital']:
                r = put('            ' + w['name'], w['amount']); style(r); last = r
    r = put('Net cash provided by/(used in) operating activities')
    ws.cell(r, 2).value = f'=SUM(B{first}:B{last})' if first else 0
    style(r, bold=True, border=rules['top_bottom'])
    sec_rows = {'operating': r}

    # Investing + Financing (shared)
    for key, short in (('investing', 'investing'), ('financing', 'financing')):
        sec = cf[key]
        r = put('CASH FLOWS FROM %s ACTIVITIES' % short.upper()); ws.cell(r, 1).font = Font(bold=True)
        rows = []
        for ln in sec['lines']:
            r = put('        ' + ln['name'], ln['amount']); style(r); rows.append(r)
        r = put('Net cash provided by/(used in) %s activities' % short)
        ws.cell(r, 2).value = f'=SUM(B{rows[0]}:B{rows[-1]})' if rows else 0
        style(r, bold=True, border=rules['top_bottom'])
        sec_rows[key] = r

    r = put('NET INCREASE/(DECREASE) IN CASH')
    ws.cell(r, 2).value = f"=B{sec_rows['operating']}+B{sec_rows['investing']}+B{sec_rows['financing']}"
    style(r, bold=True, size=12, border=rules['double_bottom'])
    net_row = r
    r = put('Cash at beginning of period', cf['cash_begin']); style(r)
    begin_row = r
    r = put('Cash at end of period')
    ws.cell(r, 2).value = f"=B{net_row}+B{begin_row}"
    style(r, bold=True, border=rules['double_bottom'])

    # Direct extras: non-cash note + reconciliation note
    if is_direct:
        if cf.get('noncash'):
            put()
            r = put('Non-cash investing and financing transactions'); ws.cell(r, 1).font = Font(italic=True)
            for n in cf['noncash']:
                r = put('        ' + n['description'], n['amount']); style(r)
        put()
        r = put('Reconciliation of net income to net cash from operating activities')
        ws.cell(r, 1).font = Font(italic=True)
        rec = cf['reconciliation']
        r = put('        Net Income (period)', rec['net_income']); style(r)
        rfirst = rlast = r
        if rec['depreciation']:
            r = put('        Add: Depreciation', rec['depreciation']); style(r); rlast = r
        for w in rec['working_capital']:
            r = put('        ' + w['name'], w['amount']); style(r); rlast = r
        r = put('Net cash from operating activities')
        ws.cell(r, 2).value = f'=SUM(B{rfirst}:B{rlast})'
        style(r, bold=True, border=rules['top_bottom'])

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 22
    ws.sheet_view.showGridLines = False

    return _xlsx_response(wb, filename)
