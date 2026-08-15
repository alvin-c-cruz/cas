"""Screen route for the Work Order Costing & Variance Report (R-07 D5)."""
import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
import pytest

from app import db
from app.bill_of_materials.models import BillOfMaterial, BillOfMaterialLine, BillOfMaterialOperation
from app.products.models import Product
from app.stock_adjustments.service import post_movement
from app.utils import ph_now
from app.work_centers.models import WorkCenter
from app.work_orders.forms import generate_wo_number
from app.work_orders.models import WorkOrder
from app.work_orders.service import (release_work_order, issue_material, start_operation,
                                     complete_operation, complete_work_order_batch,
                                     force_close_work_order)

pytestmark = [pytest.mark.integration]


def _login(client, u):
    with client.session_transaction() as s:
        s['_user_id'] = str(u.id); s['_fresh'] = True


def _ready_wo(main_branch, accountant_user, qty_to_produce='10',
              costing_method='moving_average', standard_cost=None,
              hourly_rate='60.00', minutes='60', code_suffix='A'):
    """Mirrors tests/unit/test_work_order_costing_report.py's own helper -- builds a
    released, material-issued, operation-completed WO ready for completion/force-close."""
    out = Product(code=f'D5V-OUT-{code_suffix}', name='Out', track_inventory=True,
                  costing_method=costing_method, standard_cost=standard_cost, is_active=True)
    comp = Product(code=f'D5V-COMP-{code_suffix}', name='Comp', track_inventory=True,
                   costing_method='moving_average', is_active=True)
    db.session.add_all([out, comp]); db.session.commit()
    bom = BillOfMaterial(product_id=out.id, manufacturing_mode='discrete')
    bom.lines.append(BillOfMaterialLine(line_number=1, component_product_id=comp.id,
                                        quantity_per=Decimal('1')))
    db.session.add(bom); db.session.commit()
    wc = WorkCenter(branch_id=main_branch.id, code=f'D5V-WC-{code_suffix}', name='Line',
                    hourly_rate=Decimal(hourly_rate))
    db.session.add(wc); db.session.commit()
    bom.operations.append(BillOfMaterialOperation(sequence_no=1, work_center_id=wc.id,
                                                   operation_name='Cut'))
    db.session.commit()

    wo = WorkOrder(wo_number=generate_wo_number(), bom_id=bom.id, branch_id=main_branch.id,
                   qty_to_produce=Decimal(qty_to_produce))
    db.session.add(wo); db.session.commit()
    release_work_order(wo, None)
    db.session.commit()

    post_movement(comp, main_branch.id, 'opening', Decimal('1000'), Decimal('5.00'),
                 'stock_adjustment', 0, 'seed', accountant_user, movement_date=date(2026, 1, 1))
    db.session.commit()
    issue_material(wo.materials[0], Decimal(qty_to_produce), accountant_user)
    db.session.commit()

    op = wo.operations[0]
    start_operation(op, accountant_user)
    db.session.commit()
    op.actual_start_at = ph_now() - timedelta(minutes=int(minutes))
    db.session.commit()
    complete_operation(op, accountant_user)
    db.session.commit()
    return wo


def test_accountant_can_view_report(client, db_session, accountant_user, main_branch):
    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance')
    assert resp.status_code == 200
    assert b'Work Order Costing' in resp.data


def test_staff_cannot_view_report(client, db_session, staff_user, main_branch):
    # staff_user (unlike accountant_user) has no branches assigned by the fixture --
    # without this, the before_request branch-validation hook redirects to /login
    # before the accountant_or_above_required gate ever runs, and the test would
    # fail for the wrong reason. See tests/integration/test_accounts_payable_views.py
    # TestAccess for the same pattern.
    staff_user.set_branches([main_branch])
    db_session.commit()
    _login(client, staff_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance', follow_redirects=True)
    assert b'do not have permission' in resp.data


def test_renders_standard_costed_moving_average_and_force_closed_rows(
        client, db_session, accountant_user, main_branch, wo_control_accounts):
    """Real-data render check: standard-costed variance row, moving-average
    muted-dash row, force-closed badge, and totals footer all appear in the
    actual rendered HTML -- not just asserted at the generator-function level."""
    wo_std = _ready_wo(main_branch, accountant_user, qty_to_produce='10',
                       costing_method='standard', standard_cost=Decimal('6.00'),
                       code_suffix='STD')
    complete_work_order_batch(wo_std, Decimal('10'), accountant_user)
    db.session.commit()

    wo_ma = _ready_wo(main_branch, accountant_user, code_suffix='MA')
    complete_work_order_batch(wo_ma, Decimal('10'), accountant_user)
    db.session.commit()

    wo_fc = _ready_wo(main_branch, accountant_user, qty_to_produce='10',
                      costing_method='standard', standard_cost=Decimal('6.00'),
                      code_suffix='FC')
    complete_work_order_batch(wo_fc, Decimal('4'), accountant_user)
    db.session.commit()
    force_close_work_order(wo_fc, 'Line breakdown, aborting remainder', accountant_user)
    db.session.commit()

    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance')
    assert resp.status_code == 200
    body = resp.data.decode('utf-8')

    assert wo_std.wo_number in body
    assert wo_ma.wo_number in body
    assert wo_fc.wo_number in body
    # standard-costed WO: material 50.00 + labor 60.00 = 110.00 actual, baseline 60.00,
    # variance +50.00 (unfavorable -- shows the "+" prefix and the unfavorable class)
    assert '₱110.00' in body
    assert '+₱50.00' in body
    assert 'variance-unfavorable' in body
    # moving-average WO: no standard baseline -- muted dash, not a bare number
    assert 'text-muted">&mdash;' in body or 'text-muted">—' in body
    # force-closed WO: badge shown next to its WO #
    assert 'badge-force-closed' in body
    assert 'Force-Closed' in body
    # totals footer present
    assert 'TOTALS' in body


def test_print_route_renders(client, db_session, accountant_user, main_branch):
    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance/print')
    assert resp.status_code == 200


def test_zero_baseline_standard_costed_wo_renders_dash_not_crash(
        client, db_session, accountant_user, main_branch, wo_control_accounts):
    """Review finding 1: a standard-costed WO whose Product.standard_cost is 0.00
    posts a completion whose standard_baseline computes to exactly Decimal('0.00')
    (not None) while variance_pct stays None (division-by-zero guard). The old
    template gated all three columns (baseline/variance/variance%) on a single
    `standard_baseline is not none` check, so it took the "show the numbers"
    branch and then crashed on `row.variance_pct > 0` (None > 0 -- Jinja
    TypeError, uncaught 500). Each column must be judged independently."""
    wo = _ready_wo(main_branch, accountant_user, qty_to_produce='10',
                   costing_method='standard', standard_cost=Decimal('0.00'), code_suffix='ZB')
    complete_work_order_batch(wo, Decimal('10'), accountant_user)
    db.session.commit()

    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance')
    assert resp.status_code == 200
    body = resp.data.decode('utf-8')
    assert wo.wo_number in body
    # standard_baseline (0.00) is NOT None -- must render as a real number, not a dash
    assert '₱0.00' in body
    # variance_amount (110.00, the whole actual cost) IS shown
    assert '+₱110.00' in body
    # variance_pct IS None (baseline is exactly zero) -- must render the muted dash,
    # independently of the other two columns both having real values
    assert 'text-muted">&mdash;' in body or 'text-muted">—' in body


def test_screen_links_to_print_and_excel_carrying_the_filter(
        client, db_session, accountant_user, main_branch):
    """The report SCREEN must link to its own print and Excel routes.

    Both routes are otherwise unreachable: every other test in this file reaches them
    by URL, which proves the route works and says nothing about whether a user can get
    there. Shipped that way once -- BUG-D5-REPORT-PRINT-EXPORT-BUTTONS-MISSING, caught
    only by the pre-merge browser pass -- so it is pinned here as a render assertion.

    The links must also carry the current filter state, or Print/Export would silently
    report a different row set than the one on screen.
    """
    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id

    resp = client.get('/reports/work-order-costing-variance'
                      '?status=all&date_from=2026-01-01&date_to=2026-12-31')
    assert resp.status_code == 200
    body = resp.data.decode('utf-8')

    assert '/reports/work-order-costing-variance/print' in body, 'no Print link on the screen'
    assert '/reports/work-order-costing-variance/export/excel' in body, 'no Excel link on the screen'
    assert '>Print<' in body and '>Export Excel<' in body

    for token in ('status=all', 'date_from=2026-01-01', 'date_to=2026-12-31'):
        assert token in body, 'print/export links dropped the active filter (%s)' % token


def test_export_excel_route_returns_xlsx(client, db_session, accountant_user, main_branch):
    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance/export/excel')
    assert resp.status_code == 200
    assert resp.headers['Content-Type'] == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _sheet_rows_by_wo(resp):
    """{wo_number: [cell values]} for every data row of the exported workbook.
    Layout (app/utils/export.py): row 1 title, row 2 headers, row 3+ data."""
    ws = openpyxl.load_workbook(io.BytesIO(resp.data)).active
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            out[row[0]] = list(row)
    return ws, out


def test_export_excel_writes_money_as_numeric_cells_not_text(
        client, db_session, accountant_user, main_branch, wo_control_accounts):
    """Money/qty columns must arrive in Excel as REAL NUMBERS, so the file is
    sum-able in a spreadsheet without a text-to-number pass.

    app/utils/export.py::format_value converts a Decimal to float (numeric cell)
    but sends anything else -- including a Python float -- through str(), which
    lands a TEXT cell that merely looks numeric. So the row dicts must hand it
    the generator's Decimals untouched, exactly as budget_variance_export_excel
    already does. A float() cast here silently produces a text-cell export that
    a status-code/Content-Type-only test cannot see.
    """
    wo_std = _ready_wo(main_branch, accountant_user, qty_to_produce='10',
                       costing_method='standard', standard_cost=Decimal('6.00'),
                       code_suffix='XLSTD')
    complete_work_order_batch(wo_std, Decimal('10'), accountant_user)
    db.session.commit()

    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance/export/excel')
    assert resp.status_code == 200

    ws, rows = _sheet_rows_by_wo(resp)
    assert ws.cell(row=1, column=1).value == 'Work Order Costing & Variance Report'
    assert [c.value for c in ws[2]] == [
        'WO #', 'Product', 'Qty Completed', 'Material Cost', 'Labor Cost', 'Actual Total',
        'Standard Cost', 'Variance', 'Variance %', 'Completed']

    def _numeric(value, label):
        """openpyxl reads a whole number back as int and a fractional one as float --
        either is a real numeric cell. A str here is the text-cell bug."""
        assert isinstance(value, (int, float)) and not isinstance(value, (str, bool)), (
            f'{label} cell was {value!r} ({type(value).__name__}), expected a numeric cell')

    row = rows[wo_std.wo_number]
    # material 50.00 + labor 60.00 = 110.00 actual; baseline 60.00; variance +50.00
    for idx, expected, label in ((3, 50, 'material'), (4, 60, 'labor'), (5, 110, 'actual total'),
                                 (6, 60, 'standard baseline'), (7, 50, 'variance')):
        _numeric(row[idx], label)
        assert row[idx] == expected, f'{label} cell was {row[idx]!r}, expected {expected}'
    _numeric(row[2], 'qty completed')


def test_export_excel_leaves_variance_columns_blank_for_moving_average_wo(
        client, db_session, accountant_user, main_branch, wo_control_accounts):
    """A moving-average WO has no standard baseline: those three columns carry
    None out of the generator and must export as an EMPTY cell, never the string
    'None' and never a misleading 0."""
    wo_ma = _ready_wo(main_branch, accountant_user, code_suffix='XLMA')
    complete_work_order_batch(wo_ma, Decimal('10'), accountant_user)
    db.session.commit()

    _login(client, accountant_user)
    with client.session_transaction() as s:
        s['selected_branch_id'] = main_branch.id
    resp = client.get('/reports/work-order-costing-variance/export/excel')
    assert resp.status_code == 200

    _ws, rows = _sheet_rows_by_wo(resp)
    row = rows[wo_ma.wo_number]
    assert row[5] == 110.00, 'actual total still exports for a moving-average WO'
    for idx, label in ((6, 'standard baseline'), (7, 'variance'), (8, 'variance %')):
        assert row[idx] in (None, ''), f'{label} cell was {row[idx]!r}, expected blank'
