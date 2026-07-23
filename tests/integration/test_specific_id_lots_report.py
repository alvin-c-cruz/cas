from datetime import datetime
from decimal import Decimal
from app import db
from app.settings import AppSettings
from app.utils.cache_helpers import clear_module_config_cache

D = Decimal


def _login(client, user, branch):
    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id); sess['_fresh'] = True
        sess['selected_branch_id'] = branch.id


def _enable_stock_adjustments():
    AppSettings.set_setting('module_enabled:inventory', '1', updated_by='t')
    AppSettings.set_setting('module_enabled:stock_adjustments', '1', updated_by='t')
    clear_module_config_cache()


def test_lots_report_shows_open_lots(client, db_session, admin_user, branch_main, product_specific_id):
    from app.stock_adjustments.models import StockLot
    _enable_stock_adjustments()
    lot = StockLot(product_id=product_specific_id.id, branch_id=branch_main.id,
                   original_qty=D('10'), remaining_qty=D('7'), unit_cost=D('9.50'),
                   received_at=datetime(2026, 1, 1), lot_reference='Job Order #99')
    db.session.add(lot); db.session.commit()

    _login(client, admin_user, branch_main)
    resp = client.get(f'/reports/specific-id-lots?product_id={product_specific_id.id}&branch_id={branch_main.id}')
    assert resp.status_code == 200
    assert b'Job Order #99' in resp.data
    assert b'9.50' in resp.data


def test_lots_report_drained_lot_shown_with_drained_treatment(client, db_session, admin_user, branch_main, product_specific_id):
    from app.stock_adjustments.models import StockLot
    _enable_stock_adjustments()
    lot = StockLot(product_id=product_specific_id.id, branch_id=branch_main.id,
                   original_qty=D('5'), remaining_qty=D('0'), unit_cost=D('3.00'),
                   received_at=datetime(2026, 1, 1), lot_reference='Batch Drained')
    db.session.add(lot); db.session.commit()

    _login(client, admin_user, branch_main)
    resp = client.get(f'/reports/specific-id-lots?product_id={product_specific_id.id}&branch_id={branch_main.id}')
    assert resp.status_code == 200
    assert b'drained' in resp.data.lower()


def test_lots_report_export_excel(client, db_session, admin_user, branch_main, product_specific_id):
    from app.stock_adjustments.models import StockLot
    _enable_stock_adjustments()
    lot = StockLot(product_id=product_specific_id.id, branch_id=branch_main.id,
                   original_qty=D('10'), remaining_qty=D('10'), unit_cost=D('4.00'),
                   received_at=datetime(2026, 1, 1), lot_reference=None)
    db.session.add(lot); db.session.commit()

    _login(client, admin_user, branch_main)
    resp = client.get(f'/reports/specific-id-lots/export/excel?product_id={product_specific_id.id}&branch_id={branch_main.id}')
    assert resp.status_code == 200
    assert resp.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def test_lots_report_reference_fallback_matches_screen_and_export(
        client, db_session, admin_user, branch_main, product_specific_id):
    # Reviewer finding: a lot with no lot_reference showed a bare '--' on screen but the
    # received date in the Excel export -- the same lot rendered two different things on its
    # two surfaces. Both must now show the same received-date fallback (matching the approved
    # mockup), proving the screen route and the export route agree for the SAME data.
    import io
    from openpyxl import load_workbook
    from app.stock_adjustments.models import StockLot
    _enable_stock_adjustments()
    lot = StockLot(product_id=product_specific_id.id, branch_id=branch_main.id,
                   original_qty=D('6'), remaining_qty=D('6'), unit_cost=D('5.25'),
                   received_at=datetime(2026, 3, 14), lot_reference=None)
    db.session.add(lot); db.session.commit()
    expected_fallback = '2026-03-14'

    _login(client, admin_user, branch_main)

    screen_resp = client.get(
        f'/reports/specific-id-lots?product_id={product_specific_id.id}&branch_id={branch_main.id}')
    assert screen_resp.status_code == 200
    # The received date must appear TWICE on screen (Reference column fallback + Received
    # Date column) -- not once, with a bare em-dash standing in for the missing reference.
    assert screen_resp.data.count(expected_fallback.encode()) == 2
    assert '<td>—</td>'.encode() not in screen_resp.data

    export_resp = client.get(
        f'/reports/specific-id-lots/export/excel?product_id={product_specific_id.id}&branch_id={branch_main.id}')
    assert export_resp.status_code == 200
    wb = load_workbook(io.BytesIO(export_resp.get_data()))
    ws = wb.active
    cell_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert expected_fallback in cell_values
