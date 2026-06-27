"""Integration tests for the Cash Disbursements Journal view and export."""
from datetime import date
from decimal import Decimal
import pytest

from app import db
from app.accounts.models import Account
from app.branches.models import Branch
from app.journal_entries.models import JournalEntry, JournalEntryLine
from app.users.models import User

pytestmark = [pytest.mark.journals, pytest.mark.integration]


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def branch(db_session):
    b = Branch(name='Test Branch', code='TST')
    db.session.add(b)
    db.session.commit()
    return b


@pytest.fixture()
def accountant(db_session, branch):
    from app.users.module_access import default_all_permissions
    u = User(username='acc_cd', email='acc_cd@test.com', full_name='CD Accountant',
             role='accountant', is_active=True)
    u.set_password('pass')
    # Accountants are now gated by book_permissions (Task 3); grant all so this
    # fixture user can reach /journals/cd (payments module) and siblings.
    u.set_book_permissions(default_all_permissions())
    db.session.add(u)
    db.session.flush()
    u.branches.append(branch)
    db.session.commit()
    return u


@pytest.fixture()
def cash_acct(db_session):
    a = Account.query.filter_by(code='10101').first()
    if not a:
        a = Account(code='10101', name='Cash on Hand', account_type='Asset',
                    normal_balance='debit', is_active=True)
        db.session.add(a)
        db.session.commit()
    return a


@pytest.fixture()
def expense_acct(db_session):
    a = Account.query.filter_by(code='60400').first()
    if not a:
        a = Account(code='60400', name='Rent Expense', account_type='Expense',
                    normal_balance='debit', is_active=True)
        db.session.add(a)
        db.session.commit()
    return a


def _login(client, username, password='pass'):
    client.post('/login', data={'username': username, 'password': password}, follow_redirects=True)
    with client.session_transaction() as s:
        from app.branches.models import Branch
        branch = Branch.query.first()
        if branch:
            s['selected_branch_id'] = branch.id


def _disbursement_je(branch_id, entry_date, number, cash_acct, expense_acct, amount):
    """Post a disbursement JE: Dr Expense, Cr Cash."""
    je = JournalEntry(
        entry_number=number, entry_date=entry_date,
        description='Test disbursement', reference=number,
        entry_type='disbursement',
        branch_id=branch_id, status='posted', is_balanced=True,
        total_debit=Decimal(str(amount)), total_credit=Decimal(str(amount)),
    )
    db.session.add(je)
    db.session.flush()
    db.session.add(JournalEntryLine(
        entry_id=je.id, line_number=1, account_id=expense_acct.id,
        debit_amount=Decimal(str(amount)), credit_amount=Decimal('0')))
    db.session.add(JournalEntryLine(
        entry_id=je.id, line_number=2, account_id=cash_acct.id,
        debit_amount=Decimal('0'), credit_amount=Decimal(str(amount))))
    db.session.commit()
    return je


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestCDJournalView:
    def test_redirects_to_login_when_unauthenticated(self, client):
        resp = client.get('/journals/cd', follow_redirects=False)
        assert resp.status_code == 302
        assert '/login' in resp.headers['Location']

    def test_returns_200_when_authenticated(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd')
        assert resp.status_code == 200
        assert b'Cash Disbursements Journal' in resp.data

    def test_empty_state_when_no_entries(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd?month=1&year=2020')
        assert resp.status_code == 200
        assert b'No CD entries found' in resp.data

    def test_shows_period_label(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd?mode=month&year=2026&month=6')
        assert resp.status_code == 200
        assert b'June 2026' in resp.data

    def test_shows_disbursement_entry(self, client, accountant, branch, cash_acct, expense_acct, db_session):
        je = _disbursement_je(branch.id, date(2026, 6, 10), 'CD-2026-06-0001',
                              cash_acct, expense_acct, 5000)
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd?mode=month&year=2026&month=6')
        assert resp.status_code == 200
        assert b'CD-2026-06-0001' in resp.data
        assert b'5,000.00' in resp.data

    def test_custom_range_filter(self, client, accountant, branch, cash_acct, expense_acct, db_session):
        _disbursement_je(branch.id, date(2026, 5, 15), 'CD-2026-05-0001',
                         cash_acct, expense_acct, 3000)
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd?mode=custom&date_from=2026-05-01&date_to=2026-05-31')
        assert resp.status_code == 200
        assert b'CD-2026-05-0001' in resp.data

    def test_entry_outside_period_not_shown(self, client, accountant, branch, cash_acct, expense_acct, db_session):
        _disbursement_je(branch.id, date(2026, 4, 1), 'CD-2026-04-0001',
                         cash_acct, expense_acct, 1000)
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd?mode=month&year=2026&month=6')
        assert b'CD-2026-04-0001' not in resp.data

    def test_shows_download_and_print_buttons(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd')
        assert b'Download Excel' in resp.data
        assert b'Print' in resp.data

    def test_print_route_returns_200(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd/print')
        assert resp.status_code == 200
        assert b'CASH DISBURSEMENTS JOURNAL' in resp.data


class TestCDJournalExport:
    def test_export_redirects_unauthenticated(self, client):
        resp = client.get('/journals/cd/export', follow_redirects=False)
        assert resp.status_code == 302
        assert '/login' in resp.headers['Location']

    def test_export_returns_xlsx(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd/export?mode=month&year=2026&month=6')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.headers['Content-Type']
        assert 'CD-Journal-' in resp.headers['Content-Disposition']

    def test_export_filename_uses_year_month(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd/export?mode=month&year=2026&month=3')
        assert '2026-03' in resp.headers['Content-Disposition']

    def test_export_filename_custom_range(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd/export?mode=custom&date_from=2026-01-01&date_to=2026-03-31')
        assert '2026-01-01' in resp.headers['Content-Disposition']

    def test_export_contains_entry_data(self, client, accountant, branch, cash_acct, expense_acct, db_session):
        from openpyxl import load_workbook
        import io
        _disbursement_je(branch.id, date(2026, 6, 5), 'CD-2026-06-0001',
                         cash_acct, expense_acct, 8000)
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd/export?mode=month&year=2026&month=6')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert 'Cash Disbursements Journal' in all_text
        assert 'CD-2026-06-0001' in all_text

    def test_export_empty_period_still_returns_xlsx(self, client, accountant, branch):
        _login(client, 'acc_cd')
        resp = client.get('/journals/cd/export?mode=month&year=2020&month=1')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.headers['Content-Type']
