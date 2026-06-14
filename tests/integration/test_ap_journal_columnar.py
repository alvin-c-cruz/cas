from decimal import Decimal
from datetime import date
import pytest

from app import db
from app.accounts.models import Account
from app.branches.models import Branch
from app.journal_entries.models import JournalEntry, JournalEntryLine
from app.journals.ap_journal_data import build_columnar
from app.users.models import User
from app.vendors.models import Vendor


def _acct(code, name, atype, normal):
    a = Account.query.filter_by(code=code).first()
    if not a:
        a = Account(code=code, name=name, account_type=atype,
                    normal_balance=normal, is_active=True)
        db.session.add(a)
        db.session.commit()
    return a


def _entry(branch_id, status, entry_date, number, lines):
    """lines: list of (account, debit, credit)."""
    je = JournalEntry(entry_number=number, entry_date=entry_date,
                      description='x', reference=number, entry_type='purchase',
                      branch_id=branch_id, status=status, is_balanced=True,
                      total_debit=Decimal('0'), total_credit=Decimal('0'))
    db.session.add(je)
    db.session.flush()
    n = 1
    for acct, dr, cr in lines:
        db.session.add(JournalEntryLine(
            entry_id=je.id, line_number=n, account_id=acct.id,
            debit_amount=Decimal(str(dr)), credit_amount=Decimal(str(cr))))
        n += 1
    je.total_debit = sum((Decimal(str(dr)) for _, dr, _ in lines), Decimal('0'))
    je.total_credit = sum((Decimal(str(cr)) for _, _, cr in lines), Decimal('0'))
    je.is_balanced = (je.total_debit == je.total_credit)
    db.session.commit()
    return je


def test_build_columnar_posted_pivot_and_balance(db_session):
    branch = Branch(name='Main', code='MAIN'); db.session.add(branch); db.session.commit()
    ap = _acct('20101', 'Accounts Payable - Trade', 'Liability', 'credit')
    wt = _acct('20301', 'WHT Payable - Expanded', 'Liability', 'credit')
    vat = _acct('10610', 'Input VAT', 'Asset', 'debit')
    rent = _acct('60400', 'Rent Expense', 'Expense', 'debit')

    # Bill: Dr Rent 10,000 + Dr Input VAT 1,200 ; Cr WHT 200 + Cr AP 11,000
    je = _entry(branch.id, 'posted', date(2026, 6, 1), 'JE-1',
                [(rent, 10000, 0), (vat, 1200, 0), (wt, 0, 200), (ap, 0, 11000)])

    matrix = build_columnar(
        posted_entries=[je], draft_entries=[],
        ap_account_id=ap.id, wt_account_id=wt.id, input_vat_account_ids={vat.id})

    codes = [c['code'] for c in matrix['columns']]
    assert codes == ['20101', '20301', '10610', '60400']  # AP, WHT, VAT, other
    row = matrix['rows'][0]
    assert row['is_draft'] is False
    assert row['cells'][ap.id] == Decimal('-11000')   # credit → negative
    assert row['cells'][rent.id] == Decimal('10000')
    assert matrix['totals'][ap.id] == Decimal('-11000')
    assert matrix['grand_total'] == Decimal('0')
    assert matrix['balanced'] is True


def test_build_columnar_draft_excluded_from_totals_and_columns(db_session):
    branch = Branch(name='B2', code='B2'); db.session.add(branch); db.session.commit()
    ap = _acct('20101', 'Accounts Payable - Trade', 'Liability', 'credit')
    rent = _acct('60400', 'Rent Expense', 'Expense', 'debit')
    util = _acct('60500', 'Utilities Expense', 'Expense', 'debit')

    posted = _entry(branch.id, 'posted', date(2026, 6, 2), 'JE-P',
                    [(rent, 5000, 0), (ap, 0, 5000)])
    draft = _entry(branch.id, 'draft', date(2026, 6, 3), 'JE-D',
                   [(util, 999, 0), (ap, 0, 999)])

    matrix = build_columnar([posted], [draft], ap.id, None, set())

    codes = [c['code'] for c in matrix['columns']]
    assert '60500' not in codes               # draft-only account makes no column
    assert matrix['totals'].get(rent.id) == Decimal('5000')
    # draft row present, flagged, no cells
    draft_rows = [r for r in matrix['rows'] if r['is_draft']]
    assert len(draft_rows) == 1
    assert draft_rows[0]['cells'] == {}


def _login(client, db_session, branch):
    u = User(username='acc', email='acc@t.com', full_name='Acc', role='accountant', is_active=True)
    u.set_password('pass')
    u.branches.append(branch)
    db.session.add(u)
    db.session.commit()
    client.post('/login', data={'username': 'acc', 'password': 'pass'}, follow_redirects=True)
    with client.session_transaction() as sess:
        sess['selected_branch_id'] = branch.id


def test_ap_journal_view_renders_account_columns(client, db_session):
    branch = Branch(name='Main', code='MAIN')
    db.session.add(branch)
    db.session.commit()
    ap = _acct('20101', 'Accounts Payable - Trade', 'Liability', 'credit')
    rent = _acct('60400', 'Rent Expense', 'Expense', 'debit')
    _entry(branch.id, 'posted', date(2026, 6, 1), 'AP-2026-06-0001',
           [(rent, 5000, 0), (ap, 0, 5000)])
    _login(client, db_session, branch)

    res = client.get('/journals/ap?mode=month&year=2026&month=6')
    assert res.status_code == 200
    body = res.get_data(as_text=True)
    assert 'Rent Expense' in body
    assert 'Accounts Payable - Trade' in body
    assert 'For the month of June 2026' in body


def test_ap_journal_export_returns_xlsx(client, db_session):
    branch = Branch(name='Main', code='MAIN')
    db.session.add(branch)
    db.session.commit()
    ap = _acct('20101', 'Accounts Payable - Trade', 'Liability', 'credit')
    rent = _acct('60400', 'Rent Expense', 'Expense', 'debit')
    _entry(branch.id, 'posted', date(2026, 6, 1), 'AP-2026-06-0001',
           [(rent, 5000, 0), (ap, 0, 5000)])
    _login(client, db_session, branch)

    res = client.get('/journals/ap/export?mode=month&year=2026&month=6')
    assert res.status_code == 200
    assert res.headers['Content-Type'].startswith('application/vnd.openxmlformats')
    assert 'AP-Journal-2026-06.xlsx' in res.headers['Content-Disposition']


def test_ap_journal_view_shows_draft_indicator(client, db_session):
    branch = Branch(name='Main', code='MAIN')
    db.session.add(branch)
    db.session.commit()
    ap = _acct('20101', 'Accounts Payable - Trade', 'Liability', 'credit')
    rent = _acct('60400', 'Rent Expense', 'Expense', 'debit')
    _entry(branch.id, 'draft', date(2026, 6, 5), 'AP-2026-06-0009',
           [(rent, 700, 0), (ap, 0, 700)])
    _login(client, db_session, branch)
    res = client.get('/journals/ap?mode=month&year=2026&month=6')
    body = res.get_data(as_text=True)
    assert 'Draft' in body


def _voided_bill(branch_id, bill_number, bill_date, vendor_name='Vendor V'):
    from app.purchase_bills.models import PurchaseBill
    from decimal import Decimal
    vendor = Vendor.query.filter_by(code='TEST-V').first()
    if not vendor:
        vendor = Vendor(code='TEST-V', name=vendor_name)
        db.session.add(vendor)
        db.session.flush()
    b = PurchaseBill(
        bill_number=bill_number,
        bill_date=bill_date,
        due_date=bill_date,
        vendor_id=vendor.id,
        vendor_name=vendor_name,
        status='voided',
        subtotal=Decimal('0'),
        vat_amount=Decimal('0'),
        withholding_tax_amount=Decimal('0'),
        total_amount=Decimal('0'),
        branch_id=branch_id,
    )
    db.session.add(b)
    db.session.commit()
    return b


def test_ap_journal_view_shows_voided_bill(client, db_session):
    branch = Branch(name='Main', code='MAIN')
    db.session.add(branch)
    db.session.commit()
    _voided_bill(branch.id, 'AP-2026-06-0003', date(2026, 6, 3))
    _login(client, db_session, branch)

    res = client.get('/journals/ap?mode=month&year=2026&month=6')
    assert res.status_code == 200
    body = res.get_data(as_text=True)
    assert 'AP-2026-06-0003' in body
    assert 'VOIDED' in body


def test_ap_journal_export_includes_voided_bill(client, db_session):
    import io
    from openpyxl import load_workbook

    branch = Branch(name='Main', code='MAIN')
    db.session.add(branch)
    db.session.commit()
    _voided_bill(branch.id, 'AP-2026-06-0007', date(2026, 6, 7), vendor_name='Void Co')
    _login(client, db_session, branch)

    res = client.get('/journals/ap/export?mode=month&year=2026&month=6')
    assert res.status_code == 200
    wb = load_workbook(io.BytesIO(res.get_data()))
    ws = wb.active
    all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert 'AP-2026-06-0007' in all_text
    assert '[VOIDED]' in all_text


def test_ap_journal_view_voided_excluded_from_totals_mixed(client, db_session):
    branch = Branch(name='Main', code='MAIN')
    db.session.add(branch); db.session.commit()
    ap = _acct('20101', 'Accounts Payable - Trade', 'Liability', 'credit')
    rent = _acct('60400', 'Rent Expense', 'Expense', 'debit')
    _entry(branch.id, 'posted', date(2026, 6, 1), 'AP-2026-06-0001',
           [(rent, 5000, 0), (ap, 0, 5000)])
    _voided_bill(branch.id, 'AP-2026-06-0002', date(2026, 6, 2))
    _login(client, db_session, branch)

    res = client.get('/journals/ap?mode=month&year=2026&month=6')
    assert res.status_code == 200
    body = res.get_data(as_text=True)
    assert 'AP-2026-06-0001' in body
    assert 'AP-2026-06-0002' in body
    assert 'VOIDED' in body
    assert '5,000.00' in body
