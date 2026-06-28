"""Integration tests for Books of Accounts Print-All and Export-All (Task 8)."""
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import db
from app.settings import AppSettings

pytestmark = [pytest.mark.integration]


def _login(client, user):
    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True


def _select_branch(client, branch_id):
    with client.session_transaction() as sess:
        sess['selected_branch_id'] = branch_id


def _set_company():
    db.session.add(AppSettings(key='company_name', value='Acme Trading Inc.'))
    db.session.commit()


def test_print_all_renders_six_sections(client, db_session, main_branch, admin_user):
    _set_company()
    _login(client, admin_user)
    _select_branch(client, main_branch.id)
    resp = client.get(
        '/reports/books-of-accounts/print-all?date_from=2026-01-01&date_to=2026-12-31'
    )
    assert resp.status_code == 200
    for t in [b'GENERAL JOURNAL', b'GENERAL LEDGER', b'SALES JOURNAL',
              b'PURCHASE JOURNAL', b'CASH RECEIPTS BOOK', b'CASH DISBURSEMENTS BOOK']:
        assert t in resp.data
    assert resp.data.count(b'page-break-before') >= 5


def test_export_all_has_six_sheets(client, db_session, main_branch, admin_user):
    _set_company()
    _login(client, admin_user)
    _select_branch(client, main_branch.id)
    resp = client.get(
        '/reports/books-of-accounts/export-all?date_from=2026-01-01&date_to=2026-12-31'
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.data))
    assert set(wb.sheetnames) == {
        'General Journal', 'General Ledger', 'Sales Journal',
        'Purchase Journal', 'Cash Receipts Book', 'Cash Disbursements Book',
    }


def test_print_all_requires_login(client, db_session):
    resp = client.get('/reports/books-of-accounts/print-all')
    assert resp.status_code in (302, 401)


def test_export_all_requires_login(client, db_session):
    resp = client.get('/reports/books-of-accounts/export-all')
    assert resp.status_code in (302, 401)


def test_print_all_requires_accountant_or_admin(client, db_session, main_branch, staff_user):
    _login(client, staff_user)
    _select_branch(client, main_branch.id)
    resp = client.get('/reports/books-of-accounts/print-all', follow_redirects=False)
    assert resp.status_code == 302


def test_export_all_requires_accountant_or_admin(client, db_session, main_branch, staff_user):
    _login(client, staff_user)
    _select_branch(client, main_branch.id)
    resp = client.get('/reports/books-of-accounts/export-all', follow_redirects=False)
    assert resp.status_code == 302


def test_print_all_flows_sale_and_purchase_amounts(client, db_session, main_branch, admin_user,
                                                    cash_account, revenue_account,
                                                    expense_account):
    """Posted sale+purchase JEs must flow into the correct columnar journal and GL sections."""
    from decimal import Decimal
    from datetime import date
    from app.journal_entries.models import JournalEntry, JournalEntryLine

    _set_company()
    _login(client, admin_user)
    _select_branch(client, main_branch.id)

    def _je(entry_type, dr_acct, cr_acct, amount, num):
        e = JournalEntry(
            entry_number=num, entry_date=date(2026, 6, 10),
            description=f'{entry_type} test',
            entry_type=entry_type, branch_id=main_branch.id,
            status='posted', total_debit=amount, total_credit=amount,
            reference=num,
        )
        db.session.add(e); db.session.flush()
        db.session.add(JournalEntryLine(
            entry_id=e.id, line_number=1, account_id=dr_acct.id,
            debit_amount=amount, credit_amount=Decimal('0.00'),
        ))
        db.session.add(JournalEntryLine(
            entry_id=e.id, line_number=2, account_id=cr_acct.id,
            debit_amount=Decimal('0.00'), credit_amount=amount,
        ))
        db.session.commit()
        return e

    # Sale JE: DR Cash 1,234.00 / CR Sales Revenue 1,234.00
    _je('sale', cash_account, revenue_account, Decimal('1234.00'), 'JE-SALE-001')
    # Purchase JE: DR Office Supplies 567.00 / CR Cash 567.00
    _je('purchase', expense_account, cash_account, Decimal('567.00'), 'JE-PURCH-001')

    resp = client.get(
        '/reports/books-of-accounts/print-all'
        '?date_from=2026-06-01&date_to=2026-06-30'
    )
    assert resp.status_code == 200
    body = resp.data

    # Sales Journal: cash_account column should appear (DR side of sale JE)
    assert b'Cash on Hand' in body, 'Cash on Hand missing from print-all (expected in Sales Journal + GL)'
    # The sale amount must be present in the output
    assert b'1,234.00' in body, 'Sale amount 1,234.00 not found in print-all output'
    # The purchase amount must be present
    assert b'567.00' in body, 'Purchase amount 567.00 not found in print-all output'
    # GL section includes both accounts' activity
    assert b'Sales Revenue' in body, 'Sales Revenue account missing from GL section'
    assert b'Office Supplies' in body, 'Office Supplies account missing from GL section'


def test_export_all_xlsx_includes_voided_ap_row(client, db_session, main_branch, accountant_user):
    """Voided AP bills must appear in the Purchase Journal sheet as [VOIDED] rows,
    matching the print template's behaviour — not silently dropped."""
    from decimal import Decimal
    from datetime import date
    from app.accounts_payable.models import AccountsPayable
    from app.vendors.models import Vendor

    _set_company()
    _login(client, accountant_user)
    _select_branch(client, main_branch.id)

    # Create a voided AP bill in the period
    vendor = Vendor(code='V001', name='Test Vendor')
    db.session.add(vendor)
    db.session.flush()
    bill = AccountsPayable(
        ap_number='AP-2026-06-9999',
        ap_date=date(2026, 6, 15),
        due_date=date(2026, 6, 15),
        vendor_id=vendor.id,
        vendor_name='Test Vendor',
        status='voided',
        subtotal=Decimal('0'),
        vat_amount=Decimal('0'),
        withholding_tax_amount=Decimal('0'),
        total_amount=Decimal('0'),
        branch_id=main_branch.id,
    )
    db.session.add(bill)
    db.session.commit()

    resp = client.get(
        '/reports/books-of-accounts/export-all?date_from=2026-06-01&date_to=2026-06-30'
    )
    assert resp.status_code == 200

    wb = load_workbook(BytesIO(resp.data))
    ws = wb['Purchase Journal']
    all_vals = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]

    # The voided bill must be present as a marker row, not dropped
    assert '[VOIDED]' in all_vals, (
        'Voided AP bill row was dropped from Export-All xlsx Purchase Journal sheet; '
        'it must appear as a [VOIDED] marker row to match the Print-All template.'
    )
    # The bill reference should also appear in the particulars column
    assert 'AP-2026-06-9999' in all_vals, (
        'Voided AP bill reference AP-2026-06-9999 missing from Purchase Journal sheet.'
    )
