"""Integration tests for the Cash Receipts Journal view and export."""
from datetime import date
from decimal import Decimal
import pytest

from app import db
from app.accounts.models import Account
from app.branches.models import Branch
from app.journal_entries.models import JournalEntry, JournalEntryLine
from app.users.models import User

pytestmark = [pytest.mark.journals, pytest.mark.integration]


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def branch(db_session):
    b = Branch(name='Test Branch CR', code='TCR')
    db.session.add(b)
    db.session.commit()
    return b


@pytest.fixture()
def accountant(db_session, branch):
    from app.users.module_access import default_all_permissions
    u = User(username='acc_cr', email='acc_cr@test.com', full_name='CR Accountant',
             role='accountant', is_active=True)
    u.set_password('pass')
    # Accountants are now gated by book_permissions (Task 3); grant all so this
    # fixture user can reach /journals/cr (collections module) and siblings.
    u.set_book_permissions(default_all_permissions())
    db.session.add(u)
    db.session.flush()
    u.branches.append(branch)
    db.session.commit()
    return u


@pytest.fixture()
def cash_acct(db_session):
    a = Account.query.filter_by(code='10101').first()
    if not a:
        a = Account(code='10101', name='Cash on Hand', account_type='Asset',
                    normal_balance='debit', is_active=True)
        db.session.add(a)
        db.session.commit()
    return a


@pytest.fixture()
def revenue_acct(db_session):
    a = Account.query.filter_by(code='40101').first()
    if not a:
        a = Account(code='40101', name='Service Revenue', account_type='Income',
                    normal_balance='credit', is_active=True)
        db.session.add(a)
        db.session.commit()
    return a


def _login(client, username, password='pass'):
    client.post('/login', data={'username': username, 'password': password}, follow_redirects=True)
    with client.session_transaction() as s:
        from app.branches.models import Branch
        branch = Branch.query.first()
        if branch:
            s['selected_branch_id'] = branch.id


def _receipt_je(branch_id, entry_date, number, cash_acct, revenue_acct, amount):
    """Post a receipt JE: Dr Cash, Cr Revenue."""
    je = JournalEntry(
        entry_number=number, entry_date=entry_date,
        description='Test receipt', reference=number,
        entry_type='receipt',
        branch_id=branch_id, status='posted', is_balanced=True,
        total_debit=Decimal(str(amount)), total_credit=Decimal(str(amount)),
    )
    db.session.add(je)
    db.session.flush()
    db.session.add(JournalEntryLine(
        entry_id=je.id, line_number=1, account_id=cash_acct.id,
        debit_amount=Decimal(str(amount)), credit_amount=Decimal('0')))
    db.session.add(JournalEntryLine(
        entry_id=je.id, line_number=2, account_id=revenue_acct.id,
        debit_amount=Decimal('0'), credit_amount=Decimal(str(amount))))
    db.session.commit()
    return je


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestCRJournalView:
    def test_redirects_to_login_when_unauthenticated(self, client):
        resp = client.get('/journals/cr', follow_redirects=False)
        assert resp.status_code == 302
        assert '/login' in resp.headers['Location']

    def test_returns_200_when_authenticated(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr')
        assert resp.status_code == 200
        assert b'Cash Receipts Journal' in resp.data

    def test_empty_state_when_no_entries(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr?month=1&year=2020')
        assert resp.status_code == 200
        assert b'No CR entries found' in resp.data

    def test_shows_period_label(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr?mode=month&year=2026&month=6')
        assert resp.status_code == 200
        assert b'June 2026' in resp.data

    def test_shows_receipt_entry_in_columnar_grid(self, client, accountant, branch,
                                                   cash_acct, revenue_acct, db_session):
        _receipt_je(branch.id, date(2026, 6, 10), 'CR-2026-06-0001',
                    cash_acct, revenue_acct, 5000)
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr?mode=month&year=2026&month=6')
        assert resp.status_code == 200
        assert b'CR-2026-06-0001' in resp.data
        assert b'5,000.00' in resp.data

    def test_custom_range_filter(self, client, accountant, branch,
                                  cash_acct, revenue_acct, db_session):
        _receipt_je(branch.id, date(2026, 5, 15), 'CR-2026-05-0001',
                    cash_acct, revenue_acct, 3000)
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr?mode=custom&date_from=2026-05-01&date_to=2026-05-31')
        assert resp.status_code == 200
        assert b'CR-2026-05-0001' in resp.data

    def test_entry_outside_period_not_shown(self, client, accountant, branch,
                                             cash_acct, revenue_acct, db_session):
        _receipt_je(branch.id, date(2026, 4, 1), 'CR-2026-04-0001',
                    cash_acct, revenue_acct, 1000)
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr?mode=month&year=2026&month=6')
        assert b'CR-2026-04-0001' not in resp.data

    def test_entry_in_different_branch_not_shown(self, client, accountant, branch,
                                                   cash_acct, revenue_acct, db_session):
        other_branch = Branch(name='Other Branch CR', code='OCR')
        db.session.add(other_branch)
        db.session.commit()
        _receipt_je(other_branch.id, date(2026, 6, 12), 'CR-2026-06-OTHER',
                    cash_acct, revenue_acct, 2000)
        _login(client, 'acc_cr')
        # _login sets session to the first branch (branch fixture), not other_branch
        resp = client.get('/journals/cr?mode=month&year=2026&month=6')
        assert b'CR-2026-06-OTHER' not in resp.data

    def test_shows_download_and_print_buttons(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr')
        assert b'Download Excel' in resp.data
        assert b'Print' in resp.data

    def test_print_route_returns_200(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr/print')
        assert resp.status_code == 200
        assert b'CASH RECEIPTS JOURNAL' in resp.data

    def test_role_gating_viewer_can_see(self, client, db_session, branch):
        from app.users.module_access import default_all_permissions
        viewer = User(username='viewer_cr', email='viewer_cr@test.com',
                      full_name='CR Viewer', role='viewer', is_active=True)
        viewer.set_password('pass')
        # Viewers are now gated by book_permissions (Task 3); grant all so this
        # viewer can reach /journals/cr (collections module).
        viewer.set_book_permissions(default_all_permissions())
        db.session.add(viewer)
        db.session.flush()
        # Viewers need explicit branch assignment to pass branch-validation hook
        viewer.branches.append(branch)
        db.session.commit()
        _login(client, 'viewer_cr')
        resp = client.get('/journals/cr')
        # Viewers with the collections permission can access the journal (read-only)
        assert resp.status_code == 200


class TestCRJournalExport:
    def test_export_redirects_unauthenticated(self, client):
        resp = client.get('/journals/cr/export', follow_redirects=False)
        assert resp.status_code == 302
        assert '/login' in resp.headers['Location']

    def test_export_returns_xlsx(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr/export?mode=month&year=2026&month=6')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.headers['Content-Type']
        assert 'CR-Journal-' in resp.headers['Content-Disposition']
        # PK magic bytes = valid zip/xlsx
        assert resp.data[:2] == b'PK'

    def test_export_filename_uses_year_month(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr/export?mode=month&year=2026&month=3')
        assert '2026-03' in resp.headers['Content-Disposition']

    def test_export_filename_custom_range(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr/export?mode=custom&date_from=2026-01-01&date_to=2026-03-31')
        assert '2026-01-01' in resp.headers['Content-Disposition']

    def test_export_contains_entry_data(self, client, accountant, branch,
                                         cash_acct, revenue_acct, db_session):
        from openpyxl import load_workbook
        import io
        _receipt_je(branch.id, date(2026, 6, 5), 'CR-2026-06-0001',
                    cash_acct, revenue_acct, 8000)
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr/export?mode=month&year=2026&month=6')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert 'Cash Receipts Journal' in all_text
        assert 'CR-2026-06-0001' in all_text

    def test_export_empty_period_still_returns_xlsx(self, client, accountant, branch):
        _login(client, 'acc_cr')
        resp = client.get('/journals/cr/export?mode=month&year=2020&month=1')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.headers['Content-Type']
        assert resp.data[:2] == b'PK'
