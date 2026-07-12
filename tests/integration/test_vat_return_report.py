import pytest
from decimal import Decimal
from datetime import date

from app import db
from app.reports.bir import get_vat_return_summary
from tests.integration.test_vat_settlement_compute import _vat_world, _je

pytestmark = [pytest.mark.integration]


@pytest.fixture(autouse=True)
def _fresh_module_cache(db_session):
    # bir_reports now defaults OFF (registry flip, chore/bir-reports-default-off), and its
    # enablement is read through an app-level SimpleCache that is NOT reset per test. A
    # sibling test that disables the module (e.g. test_chief_accountant) can leave a stale
    # '0' cached, 404-ing these gated routes. Explicitly enable it and clear the cache
    # before/after so every test in this file reads the module state fresh as ON.
    from app.settings import AppSettings
    from app.utils.cache_helpers import clear_module_config_cache
    AppSettings.set_setting('module_enabled:bir_reports', '1', updated_by='test')
    clear_module_config_cache()
    yield
    clear_module_config_cache()


def test_vat_return_summary_payable(db_session, main_branch):
    w = _vat_world(main_branch)
    _je(main_branch.id, date(2025, 7, 10), [(w['ar'].id, 120000, 0), (w['out'].id, 0, 120000)])
    _je(main_branch.id, date(2025, 8, 10), [(w['inp'].id, 50000, 0), (w['ap'].id, 0, 50000)])
    db.session.commit()
    r = get_vat_return_summary(2025, 3)
    assert r['output_vat'] == Decimal('120000.00')
    assert r['input_vat'] == Decimal('50000.00')
    assert r['net_payable'] == Decimal('70000.00')
    assert r['new_carryover'] == Decimal('0.00')
    assert 'error' not in r


def test_vat_return_summary_for_settled_quarter_uses_snapshot(db_session, main_branch, admin_user):
    from app.vat_settlement import service
    w = _vat_world(main_branch)
    _je(main_branch.id, date(2025, 7, 10), [(w['ar'].id, 120000, 0), (w['out'].id, 0, 120000)])
    _je(main_branch.id, date(2025, 8, 10), [(w['inp'].id, 50000, 0), (w['ap'].id, 0, 50000)])
    db.session.commit()
    service.settle_quarter(2025, 3, admin_user.id); db.session.commit()
    r = get_vat_return_summary(2025, 3)
    assert 'error' not in r
    assert r['net_payable'] == Decimal('70000.00')
    assert r['output_vat'] == Decimal('120000.00') and r['input_vat'] == Decimal('50000.00')


def test_vat_return_page_renders(client, db_session, main_branch, admin_user):
    _vat_world(main_branch); db.session.commit()
    client.post('/login', data={'username': 'admin', 'password': 'admin123'},
                follow_redirects=True)
    resp = client.get('/reports/bir/vat-return?year=2025&quarter=3')
    assert resp.status_code == 200
    assert b'VAT' in resp.data


# --- Phase 2: 2550Q schedules + reconciliation -----------------------------

def _posted_crv_regular(branch, cash_acct, rev_acct):
    """A posted CRV with one regular-VAT revenue line (1200 output VAT)."""
    from app.customers.models import Customer
    from app.cash_receipts.models import CashReceiptVoucher, CRVRevenueLine
    cust = Customer(code='VR-2550', name='Cash Customer', tin='111-222-333-000',
                    address='1 Rizal St')
    db.session.add(cust); db.session.commit()
    crv = CashReceiptVoucher(
        branch_id=branch.id, crv_number='CRV-2550-0001', crv_date=date(2025, 8, 10),
        customer_id=cust.id, customer_name=cust.name, customer_tin=cust.tin,
        cash_account_id=cash_acct.id, status='posted')
    crv.revenue_lines.append(CRVRevenueLine(
        line_number=1, description='cash sale', amount=Decimal('11200.00'),
        vat_rate=Decimal('12.00'), vat_category='V12', vat_nature='regular',
        line_total=Decimal('11200.00'), vat_amount=Decimal('1200.00'),
        account_id=rev_acct.id))
    db.session.add(crv); db.session.commit()
    return crv


def test_summary_gains_schedules_and_reconciliation(db_session, main_branch,
                                                    cash_account, revenue_account):
    _posted_crv_regular(main_branch, cash_account, revenue_account)
    r = get_vat_return_summary(2025, 3)
    assert 'sales_schedule' in r and 'input_schedule' in r and 'reconciliation' in r
    row_12a = next(x for x in r['sales_schedule']['rows'] if x['box'] == '12A')
    assert row_12a['tax'] == Decimal('1200.00')
    assert r['reconciliation']['output_docs'] == Decimal('1200.00')


def test_summary_reconciliation_gl_unavailable_on_error(db_session, main_branch):
    # No VAT accounts / settlement settings -> compute_vat_position raises ->
    # summary carries an error, GL side is None, but schedules still render.
    r = get_vat_return_summary(2025, 3)
    assert r['reconciliation']['output_gl'] is None
    assert r['reconciliation']['in_balance'] is False
    assert 'sales_schedule' in r


def test_print_facsimile_renders_box_numbers_and_footnotes(
        client, db_session, main_branch, cash_account, revenue_account, admin_user):
    _posted_crv_regular(main_branch, cash_account, revenue_account)
    client.post('/login', data={'username': 'admin', 'password': 'admin123'},
                follow_redirects=True)
    resp = client.get('/reports/bir/vat-return/print?year=2025&quarter=3')
    assert resp.status_code == 200
    for box in (b'12A', b'12D', b'18A', b'18G', b'23', b'26'):
        assert box in resp.data, box
    # inline footnotes (spec 5): item 23 zero, capital-goods not split
    assert b'Creditable VAT Withheld' in resp.data
    assert b'1,000,000' in resp.data


def test_page_shows_schedules_and_reconciliation_banner(
        client, db_session, main_branch, cash_account, revenue_account, admin_user):
    _posted_crv_regular(main_branch, cash_account, revenue_account)
    client.post('/login', data={'username': 'admin', 'password': 'admin123'},
                follow_redirects=True)
    resp = client.get('/reports/bir/vat-return?year=2025&quarter=3')
    assert resp.status_code == 200
    assert b'Vatable Sales' in resp.data          # Part I schedule row
    assert b'12A' in resp.data                     # box number rendered
    assert b'Capital Goods' in resp.data           # Part II schedule row


def test_page_shows_out_of_balance_banner(client, db_session, main_branch, admin_user):
    # raw-JE world posts output VAT with no document lines -> docs != GL
    w = _vat_world(main_branch)
    _je(main_branch.id, date(2025, 7, 10), [(w['ar'].id, 120000, 0), (w['out'].id, 0, 120000)])
    db.session.commit()
    client.post('/login', data={'username': 'admin', 'password': 'admin123'},
                follow_redirects=True)
    resp = client.get('/reports/bir/vat-return?year=2025&quarter=3')
    assert resp.status_code == 200
    assert b'do not tie to' in resp.data           # reconciliation banner fired


def test_excel_export_includes_schedule_rows(
        client, db_session, main_branch, cash_account, revenue_account, admin_user):
    import io
    from openpyxl import load_workbook
    _posted_crv_regular(main_branch, cash_account, revenue_account)
    client.post('/login', data={'username': 'admin', 'password': 'admin123'},
                follow_redirects=True)
    resp = client.get('/reports/bir/vat-return/export/excel?year=2025&quarter=3')
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    text = '\n'.join(str(c.value) for row in wb.active.iter_rows() for c in row if c.value)
    assert 'Vatable Sales' in text          # a Part I schedule row is present
    assert 'Capital Goods' in text          # a Part II schedule row is present
    assert '1200.00' in text
