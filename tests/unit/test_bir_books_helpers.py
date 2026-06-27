from openpyxl import Workbook
from app.settings import AppSettings
from app import db
from app.utils.bir_books import get_company_identity, write_bir_book_header


def _set(key, value):
    db.session.add(AppSettings(key=key, value=value)); db.session.commit()


def test_get_company_identity_reads_settings(db_session):
    _set('company_name', 'Acme Trading Inc.'); _set('company_tin', '123-456-789')
    _set('tin_branch_code', '00000'); _set('rdo_code', '050')
    _set('company_address', '1 Rizal St, Manila')
    assert get_company_identity() == {
        'name': 'Acme Trading Inc.', 'tin': '123-456-789', 'tin_branch': '00000',
        'rdo': '050', 'address': '1 Rizal St, Manila'}


def test_get_company_identity_blank_when_unset(db_session):
    ident = get_company_identity()
    assert ident['name'] == '' and ident['tin'] == '' and ident['rdo'] == ''


def test_write_bir_book_header_writes_name_title_period(db_session):
    wb = Workbook(); ws = wb.active
    company = {'name': 'Acme Trading Inc.', 'tin': '123', 'tin_branch': '00000',
               'rdo': '050', 'address': '1 Rizal St'}
    last = write_bir_book_header(ws, company, 'GENERAL JOURNAL', 'For June 2026')
    assert ws.cell(1, 1).value == 'ACME TRADING INC.'
    cells = [ws.cell(r, 1).value for r in range(1, last + 1)]
    assert 'GENERAL JOURNAL' in cells and 'For June 2026' in cells
    assert any(c and c.startswith('TIN: 123-00000') for c in cells)
    assert any(c and 'RDO: 050' in c for c in cells)
