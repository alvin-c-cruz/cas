"""Tests for Income Statement (and Balance Sheet) Excel exporters.

These verify that the xlsx builders consume the new type-driven section
shapes emitted by generate_income_statement / generate_balance_sheet.
"""
import pytest
from io import BytesIO
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from app import db
from app.accounts.models import Account
from app.branches.models import Branch
from app.journal_entries.models import JournalEntry, JournalEntryLine
from app.reports.financial import generate_income_statement, generate_balance_sheet
from app.reports.statement_export import (
    build_income_statement_xlsx,
    build_balance_sheet_xlsx,
    income_statement_lines,
)

pytestmark = [pytest.mark.unit]


# ── shared helpers (mirror test_income_statement_generator.py) ────────────────

def _branch():
    b = Branch(name='Main', code='MAIN')
    db.session.add(b)
    db.session.commit()
    return b


def _acct(code, name, atype, normal='debit', parent_id=None, classification=None):
    a = Account(code=code, name=name, account_type=atype, normal_balance=normal,
                is_active=True, parent_id=parent_id, classification=classification)
    db.session.add(a)
    db.session.commit()
    return a


def _je(branch_id, lines, number):
    je = JournalEntry(entry_number=number, entry_date=date(2026, 6, 10), description='d',
                      reference=number, entry_type='adjustment', branch_id=branch_id,
                      status='posted', is_balanced=True, total_debit=Decimal('0'),
                      total_credit=Decimal('0'))
    db.session.add(je)
    db.session.flush()
    n = 1
    for acct, dr, cr in lines:
        db.session.add(JournalEntryLine(entry_id=je.id, line_number=n, account_id=acct.id,
                                        debit_amount=Decimal(str(dr)),
                                        credit_amount=Decimal(str(cr))))
        n += 1
    db.session.commit()
    return je


def _full_pl(db_session):
    """Seed a full P&L scenario — matches test_income_statement_generator._full_pl."""
    b = _branch()
    cash  = _acct('10101', 'Cash', 'Asset', 'debit', classification='Current')
    sales = _acct('40101', 'Sales - Goods', 'Revenue', 'credit')
    disc  = _acct('40104', 'Sales Discounts', 'Contra-Revenue', 'debit')
    cogs  = _acct('50101', 'Cost of Goods Sold', 'Cost of Goods Sold', 'debit')
    sell  = _acct('50211', 'Sales Commissions', 'Selling Expense', 'debit')
    admin = _acct('50221', 'Office Salaries', 'Administrative Expense', 'debit')
    oinc  = _acct('40201', 'Interest Income', 'Other Income', 'credit')
    oexp  = _acct('50301', 'Interest Expense', 'Other Expense', 'debit')
    tax   = _acct('50401', 'Income Tax - Current', 'Income Tax Expense', 'debit')
    _je(b.id, [(cash, 1000, 0), (sales, 0, 1000)], 'JE-S')
    _je(b.id, [(disc, 100, 0), (cash, 0, 100)], 'JE-D')
    _je(b.id, [(cogs, 400, 0), (cash, 0, 400)], 'JE-C')
    _je(b.id, [(sell, 50, 0), (cash, 0, 50)], 'JE-SE')
    _je(b.id, [(admin, 150, 0), (cash, 0, 150)], 'JE-AE')
    _je(b.id, [(cash, 30, 0), (oinc, 0, 30)], 'JE-OI')
    _je(b.id, [(oexp, 20, 0), (cash, 0, 20)], 'JE-OE')
    _je(b.id, [(tax, 60, 0), (cash, 0, 60)], 'JE-T')
    return b


def _xlsx_cells(xlsx_or_response):
    """Accept either raw bytes or a Flask/pytest-flask response object."""
    if hasattr(xlsx_or_response, 'data'):
        raw = xlsx_or_response.data
    else:
        raw = xlsx_or_response
    wb = load_workbook(BytesIO(raw))
    return [c.value for row in wb.active.iter_rows() for c in row if c.value is not None]


# ── Income Statement export tests ─────────────────────────────────────────────

def test_is_xlsx_has_subtotal_rows(db_session):
    """Workbook must open and contain the Net Income label."""
    data = generate_income_statement(date(2026, 6, 1), date(2026, 6, 30), None)
    xlsx = build_income_statement_xlsx(data, period_label='June 2026',
                                       company={}, branch_name=None,
                                       filename='test.xlsx')
    cells = _xlsx_cells(xlsx)
    assert 'Net Income' in cells


def test_is_xlsx_section_labels_and_subtotals(db_session):
    """Real data workbook: section labels and subtotal rows appear."""
    b = _full_pl(db_session)
    data = generate_income_statement(date(2026, 6, 1), date(2026, 6, 30), b.id)
    xlsx = build_income_statement_xlsx(data, period_label='June 2026',
                                       company={'name': 'Test Co'},
                                       branch_name='Main',
                                       filename='test.xlsx')
    cells = _xlsx_cells(xlsx)
    # Section labels
    assert 'Cost of Goods Sold' in cells, f'Expected Cost of Goods Sold in {cells}'
    assert 'Selling Expenses' in cells
    assert 'Administrative Expenses' in cells
    # Subtotal labels emitted by the generator
    assert 'Gross Profit' in cells
    assert 'Operating Income' in cells
    assert 'Net Income' in cells
    # Account line appears
    assert any('Sales - Goods' in str(c) for c in cells), f'Missing sales account in {cells}'


def test_is_xlsx_numeric_values_present(db_session):
    """Net income value (250.0) is written into a cell."""
    b = _full_pl(db_session)
    data = generate_income_statement(date(2026, 6, 1), date(2026, 6, 30), b.id)
    xlsx = build_income_statement_xlsx(data, period_label='June 2026',
                                       company={}, branch_name=None,
                                       filename='test.xlsx')
    cells = _xlsx_cells(xlsx)
    # net_income = 250.0 (from fixture)
    assert 250.0 in cells or 250 in cells, f'Net income 250 not found in cells: {cells}'


# ── Balance Sheet export tests ────────────────────────────────────────────────

def _full_bs(db_session):
    """Minimal BS seed: cash asset + equity capital."""
    b = _branch()
    cash  = _acct('10101', 'Cash on Hand', 'Asset', 'debit', classification='Current')
    cap   = _acct('30101', 'Capital Stock', 'Equity', 'credit', classification=None)
    _je(b.id, [(cash, 5000, 0), (cap, 0, 5000)], 'JE-BS')
    return b


def test_bs_xlsx_division_and_total_labels(db_session):
    """Balance Sheet workbook contains division label and TOTAL ASSETS."""
    b = _full_bs(db_session)
    bs = generate_balance_sheet(date(2026, 6, 30), branch_id=b.id)
    xlsx = build_balance_sheet_xlsx(bs, as_of_label='As of June 30, 2026',
                                    company={'name': 'Test Co'},
                                    branch_name='Main',
                                    filename='test_bs.xlsx')
    cells = _xlsx_cells(xlsx)
    assert 'TOTAL ASSETS' in cells, f'TOTAL ASSETS missing from {cells}'
    assert 'TOTAL LIABILITIES AND EQUITY' in cells
    # Account label for the cash account should appear (single non-empty division —
    # sub-header is suppressed when only one non-empty division exists)
    assert any('Cash on Hand' in str(c) for c in cells), f'Cash account missing from {cells}'


# ── Net Income print-styling contract ─────────────────────────────────────────

def test_income_statement_lines_net_income_kind_and_rule(db_session):
    """Net Income row must carry kind='net' and rule='double_bottom' for print CSS.

    The print template styles the Net Income row via `.net td { font-size: 14px; }`
    and `.rule-double_bottom td { border-bottom: 3px double #000; }`.  Emitting a
    generic subtotal kind/rule would silently drop that formatting.
    """
    b = _full_pl(db_session)
    stmt = generate_income_statement(date(2026, 6, 1), date(2026, 6, 30), b.id)
    lines = income_statement_lines(stmt)

    net_income_rows = [ln for ln in lines if ln.get('label') == 'Net Income']
    assert net_income_rows, 'No Net Income row found in income_statement_lines output'
    ni = net_income_rows[0]
    assert ni['kind'] == 'net', (
        f"Net Income row must have kind='net' for print CSS, got kind={ni['kind']!r}"
    )
    assert ni['rule'] == 'double_bottom', (
        f"Net Income row must have rule='double_bottom' for print CSS, got rule={ni['rule']!r}"
    )
