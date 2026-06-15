import io
from datetime import date, datetime
from decimal import Decimal
from unittest.mock import MagicMock

from openpyxl import load_workbook

import pytest
from app.journals.ap_journal_data import resolve_period, _fmt, build_ap_journal_xlsx, build_columnar
pytestmark = [pytest.mark.journals, pytest.mark.unit]


def test_resolve_period_defaults_to_given_month():
    p = resolve_period({}, today=date(2026, 6, 13))
    assert p['mode'] == 'month'
    assert p['date_from'] == date(2026, 6, 1)
    assert p['date_to'] == date(2026, 6, 30)
    assert p['label'] == 'For the month of June 2026'


def test_resolve_period_explicit_month():
    p = resolve_period({'mode': 'month', 'year': '2026', 'month': '2'}, today=date(2026, 6, 13))
    assert p['date_from'] == date(2026, 2, 1)
    assert p['date_to'] == date(2026, 2, 28)  # 2026 not a leap year
    assert p['label'] == 'For the month of February 2026'


def test_resolve_period_custom_range():
    p = resolve_period(
        {'mode': 'custom', 'date_from': '2026-01-15', 'date_to': '2026-03-10'},
        today=date(2026, 6, 13),
    )
    assert p['mode'] == 'custom'
    assert p['date_from'] == date(2026, 1, 15)
    assert p['date_to'] == date(2026, 3, 10)
    assert p['label'] == 'From January 15, 2026 to March 10, 2026'  # no leading zero on day


def test_resolve_period_custom_with_bad_dates_falls_back_to_month():
    p = resolve_period({'mode': 'custom', 'date_from': 'bad', 'date_to': ''}, today=date(2026, 6, 13))
    assert p['mode'] == 'month'
    assert p['date_from'] == date(2026, 6, 1)


def test_resolve_period_custom_label_no_leading_zero():
    p = resolve_period(
        {'mode': 'custom', 'date_from': '2026-03-05', 'date_to': '2026-03-09'},
        today=date(2026, 6, 13),
    )
    assert p['label'] == 'From March 5, 2026 to March 9, 2026'


def test_resolve_period_custom_inverted_falls_back_to_month():
    p = resolve_period(
        {'mode': 'custom', 'date_from': '2026-06-30', 'date_to': '2026-01-01'},
        today=date(2026, 6, 13),
    )
    assert p['mode'] == 'month'
    assert p['date_from'] == date(2026, 6, 1)


def test__fmt():
    assert _fmt(None) == ''
    assert _fmt(Decimal('0')) == ''
    assert _fmt(Decimal('5000')) == '5,000.00'
    assert _fmt(Decimal('-5000')) == '(5,000.00)'
    assert _fmt(Decimal('-0.01')) == '(0.01)'


def _fake_entry(date_str, number, invoice, vendor, notes):
    class E: pass
    e = E()
    e.entry_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    e.reference = number
    e._invoice = invoice
    e._vendor = vendor
    e._notes = notes
    return e


def test_build_ap_journal_xlsx_has_headers_and_total_row(app):
    columns = [
        {'account_id': 1, 'code': '20101', 'name': 'Accounts Payable - Trade', 'group': 'ap'},
        {'account_id': 2, 'code': '60400', 'name': 'Rent Expense', 'group': 'other'},
    ]
    rows = [{
        'entry': _fake_entry('2026-06-01', 'AP-2026-06-0001', 'SI-1', 'Vendor A', 'Rent'),
        'cells': {1: Decimal('-5000'), 2: Decimal('5000')},
        'is_draft': False,
        'is_voided': False,
    }]
    totals = {1: Decimal('-5000'), 2: Decimal('5000')}
    with app.app_context():
        resp = build_ap_journal_xlsx(
            columns=columns, rows=rows, totals=totals,
            period_label='For the month of June 2026',
            company_name='ABC Company', branch_name='Main Branch',
            filename='AP-Journal-2026-06.xlsx',
            identity=lambda e: (e.reference, e._invoice, e._vendor, e._notes))
    assert resp.headers['Content-Type'].startswith('application/vnd.openxmlformats')
    assert 'AP-Journal-2026-06.xlsx' in resp.headers['Content-Disposition']

    wb = load_workbook(io.BytesIO(resp.get_data()))
    ws = wb.active
    all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert 'Accounts Payable Journal' in all_text
    assert 'Accounts Payable - Trade' in all_text
    assert 'Rent Expense' in all_text
    assert 'TOTAL' in all_text

    fixed = ['Date', 'AP No.', 'Invoice No.', 'Vendor', 'Particulars']
    header = fixed + [c['name'] for c in columns]

    # header row 6 (branch shown in test), data row 7
    data_row = [ws.cell(row=7, column=i).value for i in range(1, len(header) + 1)]
    assert -5000.0 in data_row   # AP column (credit → negative float)
    assert 5000.0 in data_row     # Rent Expense column (debit → positive float)

    # blank row 8, TOTAL row 9 — amounts are SUM formulas (openpyxl stores as strings)
    total_row = [ws.cell(row=9, column=i).value for i in range(1, len(header) + 1)]
    assert total_row[0] == 'TOTAL'
    assert total_row[5] == '=SUM(F7:F8)'   # AP column formula
    assert total_row[6] == '=SUM(G7:G8)'   # Rent Expense column formula


def _mock_bill(ap_number, ap_date, vendor_name='Vendor X',
               vendor_invoice_number='INV-1', notes=''):
    b = MagicMock()
    b.ap_number = ap_number
    b.ap_date = ap_date
    b.vendor_name = vendor_name
    b.vendor_invoice_number = vendor_invoice_number
    b.notes = notes
    return b


def test_build_columnar_voided_rows_excluded_from_totals():
    voided = _mock_bill('AP-2026-06-0002', date(2026, 6, 2))
    matrix = build_columnar(
        posted_entries=[], draft_entries=[],
        ap_account_id=1, wt_account_id=2, input_vat_account_ids=set(),
        voided_bills=[voided],
    )
    assert len(matrix['rows']) == 1
    row = matrix['rows'][0]
    assert row['is_voided'] is True
    assert row['cells'] == {}
    assert row['ap'] is voided
    assert matrix['totals'] == {}
    assert matrix['grand_total'] == Decimal('0')


def test_build_columnar_voided_rows_sort_with_posted_by_date():
    voided = _mock_bill('AP-2026-06-0001', date(2026, 6, 1))

    posted_je = MagicMock()
    posted_je.entry_date = date(2026, 6, 3)
    posted_je.entry_number = 'AP-2026-06-0003'
    line = MagicMock()
    acct = MagicMock()
    acct.id = 99
    acct.code = '20101'
    acct.name = 'AP'
    line.account = acct
    line.debit_amount = Decimal('0')
    line.credit_amount = Decimal('5000')
    posted_je.lines.all.return_value = [line]

    matrix = build_columnar(
        posted_entries=[posted_je], draft_entries=[],
        ap_account_id=99, wt_account_id=None, input_vat_account_ids=set(),
        voided_bills=[voided],
    )
    assert matrix['rows'][0]['is_voided'] is True
    assert matrix['rows'][1]['is_voided'] is False


def test_build_ap_journal_xlsx_voided_row_has_red_fill_and_no_amounts(app):
    from datetime import date
    from decimal import Decimal
    from unittest.mock import MagicMock

    bill = MagicMock()
    bill.ap_date = date(2026, 6, 3)
    bill.ap_number = 'AP-2026-06-0002'
    bill.vendor_invoice_number = 'INV-99'
    bill.vendor_name = 'Voided Vendor'
    bill.notes = 'Test void'

    columns = [
        {'account_id': 1, 'code': '20101', 'name': 'Accounts Payable - Trade', 'group': 'ap'},
    ]
    rows = [{
        'entry': None,
        'ap': bill,
        'cells': {},
        'is_draft': False,
        'is_voided': True,
    }]
    with app.app_context():
        resp = build_ap_journal_xlsx(
            columns=columns, rows=rows, totals={},
            period_label='For the month of June 2026',
            company_name='ABC Co', branch_name=None,
            filename='test.xlsx',
            identity=lambda e: ('', '', '', ''))

    wb = load_workbook(io.BytesIO(resp.get_data()))
    ws = wb.active
    all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert 'AP-2026-06-0002' in all_text
    assert 'Voided Vendor' in all_text
    assert '[VOIDED]' in all_text

    # All amount cells for the voided row must be blank (None)
    # No branch → header row 5, data row 6
    data_row_vals = [ws.cell(row=6, column=i).value for i in range(6, 6 + len(columns))]
    assert all(v is None for v in data_row_vals)

    # Verify red fill on all cells of the voided row (data row 6, no branch → header row 5)
    # 5 fixed columns (Date, AP No., Invoice No., Vendor, Particulars) + N account columns
    for col_idx in range(1, 5 + len(columns) + 1):
        cell = ws.cell(row=6, column=col_idx)
        assert cell.fill.fgColor.rgb.endswith('FFCDD2'), f"col {col_idx}: expected FFCDD2, got {cell.fill.fgColor.rgb}"


def test_build_columnar_voided_no_column_contribution():
    voided = _mock_bill('AP-2026-06-0005', date(2026, 6, 5))
    matrix = build_columnar(
        posted_entries=[], draft_entries=[],
        ap_account_id=1, wt_account_id=2, input_vat_account_ids=set(),
        voided_bills=[voided],
    )
    assert matrix['columns'] == []
