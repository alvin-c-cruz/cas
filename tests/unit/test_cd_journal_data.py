import io
from datetime import date
from decimal import Decimal
from unittest.mock import MagicMock

from openpyxl import load_workbook
import pytest

from app.journals.cd_journal_data import build_columnar_cd, build_cd_journal_xlsx

pytestmark = [pytest.mark.journals, pytest.mark.unit]


def _mock_line(acct_id, code, name, debit, credit):
    line = MagicMock()
    acct = MagicMock()
    acct.id = acct_id
    acct.code = code
    acct.name = name
    line.account = acct
    line.debit_amount = Decimal(str(debit))
    line.credit_amount = Decimal(str(credit))
    return line


def _mock_entry(date_str, number, lines):
    je = MagicMock()
    je.entry_date = date.fromisoformat(date_str)
    je.entry_number = number
    je.reference = number
    je.lines.all.return_value = lines
    return je


# AP=1, WHT=2, VAT=3, Cash=10, Expense=20
AP_ID = 1
WHT_ID = 2
VAT_ID = 3
CASH_ID = 10
EXPENSE_ID = 20


def _standard_cdv_lines():
    """Dr Expense 10000, Dr Input VAT 1200, Cr WHT 200, Cr AP 0, Cr Cash 11000."""
    return [
        _mock_line(EXPENSE_ID, '60400', 'Rent Expense', 10000, 0),
        _mock_line(VAT_ID,     '10610', 'Input VAT',    1200,  0),
        _mock_line(WHT_ID,     '20301', 'WHT Payable',  0,     200),
        _mock_line(CASH_ID,    '10101', 'Cash on Hand', 0,     11000),
    ]


def test_build_columnar_cd_basic_pivot_and_balance():
    je = _mock_entry('2026-06-01', 'CD-2026-06-0001', _standard_cdv_lines())
    matrix = build_columnar_cd(
        posted_entries=[je], draft_entries=[],
        ap_account_id=AP_ID, wt_account_id=WHT_ID,
        input_vat_account_ids={VAT_ID},
    )
    assert matrix['balanced'] is True
    assert matrix['grand_total'] == Decimal('0')
    assert len(matrix['rows']) == 1
    row = matrix['rows'][0]
    assert row['is_draft'] is False
    assert row['is_cancelled'] is False
    # Expense debit: signed = 10000 - 0 = +10000
    assert row['cells'][EXPENSE_ID] == Decimal('10000')
    # Cash credit: signed = 0 - 11000 = -11000
    assert row['cells'][CASH_ID] == Decimal('-11000')


def test_build_columnar_cd_column_ordering():
    """Column order: ap_applied(0), vat(1), expense(2), wht(3), cash(4)."""
    je = _mock_entry('2026-06-01', 'CD-2026-06-0001', _standard_cdv_lines())
    matrix = build_columnar_cd(
        posted_entries=[je], draft_entries=[],
        ap_account_id=AP_ID, wt_account_id=WHT_ID,
        input_vat_account_ids={VAT_ID},
    )
    groups = [c['group'] for c in matrix['columns']]
    # expense before wht and cash; vat before expense
    assert groups.index('vat') < groups.index('expense')
    assert groups.index('expense') < groups.index('wht')
    assert groups.index('wht') < groups.index('cash')


def test_build_columnar_cd_column_groups():
    je = _mock_entry('2026-06-01', 'CD-2026-06-0001', _standard_cdv_lines())
    matrix = build_columnar_cd(
        posted_entries=[je], draft_entries=[],
        ap_account_id=AP_ID, wt_account_id=WHT_ID,
        input_vat_account_ids={VAT_ID},
    )
    by_id = {c['account_id']: c['group'] for c in matrix['columns']}
    assert by_id[VAT_ID]     == 'vat'
    assert by_id[EXPENSE_ID] == 'expense'
    assert by_id[WHT_ID]     == 'wht'
    assert by_id[CASH_ID]    == 'cash'


def test_build_columnar_cd_cancelled_ref_is_flagged():
    je = _mock_entry('2026-06-01', 'CD-2026-06-0001', _standard_cdv_lines())
    matrix = build_columnar_cd(
        posted_entries=[je], draft_entries=[],
        ap_account_id=AP_ID, wt_account_id=WHT_ID,
        input_vat_account_ids={VAT_ID},
        cancelled_refs={'CD-2026-06-0001'},
    )
    assert matrix['rows'][0]['is_cancelled'] is True
    # Cancelled rows are still included in totals
    assert matrix['totals'][CASH_ID] == Decimal('-11000')


def test_build_columnar_cd_cancelled_ref_not_in_set_is_not_flagged():
    je = _mock_entry('2026-06-01', 'CD-2026-06-0001', _standard_cdv_lines())
    matrix = build_columnar_cd(
        posted_entries=[je], draft_entries=[],
        ap_account_id=AP_ID, wt_account_id=WHT_ID,
        input_vat_account_ids={VAT_ID},
        cancelled_refs={'CD-2026-06-9999'},
    )
    assert matrix['rows'][0]['is_cancelled'] is False


def test_build_columnar_cd_draft_has_no_cells():
    je = _mock_entry('2026-06-01', 'CD-2026-06-0001', _standard_cdv_lines())
    draft = _mock_entry('2026-06-02', 'CD-2026-06-0002', _standard_cdv_lines())
    matrix = build_columnar_cd(
        posted_entries=[je], draft_entries=[draft],
        ap_account_id=AP_ID, wt_account_id=WHT_ID,
        input_vat_account_ids={VAT_ID},
    )
    draft_row = next(r for r in matrix['rows'] if r['is_draft'])
    assert draft_row['cells'] == {}


def test_build_cd_journal_xlsx_has_headers_and_total_row(app):
    columns = [
        {'account_id': EXPENSE_ID, 'code': '60400', 'name': 'Rent Expense', 'group': 'expense'},
        {'account_id': CASH_ID,    'code': '10101', 'name': 'Cash on Hand', 'group': 'cash'},
    ]
    rows = [{
        'entry': _mock_entry('2026-06-01', 'CD-2026-06-0001', []),
        'cells': {EXPENSE_ID: Decimal('10000'), CASH_ID: Decimal('-10000')},
        'is_draft': False,
        'is_cancelled': False,
    }]
    totals = {EXPENSE_ID: Decimal('10000'), CASH_ID: Decimal('-10000')}
    with app.app_context():
        resp = build_cd_journal_xlsx(
            columns=columns, rows=rows, totals=totals,
            period_label='For the month of June 2026',
            company_name='ABC Company', branch_name='Main Branch',
            filename='CD-Journal-2026-06.xlsx',
            identity=lambda e: ('CD-2026-06-0001', '', 'Vendor A', 'Rent'),
        )
    assert resp.headers['Content-Type'].startswith('application/vnd.openxmlformats')
    assert 'CD-Journal-2026-06.xlsx' in resp.headers['Content-Disposition']
    wb = load_workbook(io.BytesIO(resp.get_data()))
    ws = wb.active
    all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert 'Cash Disbursements Journal' in all_text
    assert 'Rent Expense' in all_text
    assert 'TOTAL' in all_text


def test_build_cd_journal_xlsx_cancelled_row_has_red_fill(app):
    columns = [
        {'account_id': CASH_ID, 'code': '10101', 'name': 'Cash on Hand', 'group': 'cash'},
    ]
    entry = _mock_entry('2026-06-01', 'CD-2026-06-0001', [])
    rows = [{
        'entry': entry,
        'cells': {CASH_ID: Decimal('-10000')},
        'is_draft': False,
        'is_cancelled': True,
    }]
    with app.app_context():
        resp = build_cd_journal_xlsx(
            columns=columns, rows=rows, totals={CASH_ID: Decimal('-10000')},
            period_label='For the month of June 2026',
            company_name='ABC Co', branch_name=None,
            filename='test.xlsx',
            identity=lambda e: ('CD-2026-06-0001', '', 'Vendor X', '[CANCELLED]'),
        )
    wb = load_workbook(io.BytesIO(resp.get_data()))
    ws = wb.active
    # No branch -> header row 5, data row 6
    for col_idx in range(1, 5 + len(columns) + 1):
        cell = ws.cell(row=6, column=col_idx)
        assert cell.fill.fgColor.rgb.endswith('FFCDD2'), \
            f"col {col_idx}: expected FFCDD2, got {cell.fill.fgColor.rgb}"
