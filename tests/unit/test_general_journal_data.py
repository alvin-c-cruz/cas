from decimal import Decimal
from datetime import date
from app import db
from app.journal_entries.models import JournalEntry, JournalEntryLine
from app.reports.general_journal_data import build_general_journal, build_general_journal_xlsx, _write_gj_rows


def _entry(branch_id, dr, cr, amount, status='posted', etype='adjustment'):
    e = JournalEntry(entry_number=f'JV-T-{dr.id}-{cr.id}-{status}', entry_date=date(2026, 6, 15),
                     description='Test adjustment', entry_type=etype, branch_id=branch_id,
                     status=status, total_debit=amount, total_credit=amount,
                     reference='JV-2026-06-0001')
    db.session.add(e); db.session.flush()
    db.session.add(JournalEntryLine(entry_id=e.id, line_number=1, account_id=dr.id,
                                    debit_amount=amount, credit_amount=Decimal('0.00')))
    db.session.add(JournalEntryLine(entry_id=e.id, line_number=2, account_id=cr.id,
                                    debit_amount=Decimal('0.00'), credit_amount=amount))
    db.session.commit()
    return e


def test_build_general_journal_splits_debits_and_credits(db_session, main_branch,
                                                         cash_account, revenue_account):
    _entry(main_branch.id, cash_account, revenue_account, Decimal('1000.00'))
    gj = build_general_journal(JournalEntry.query.all())
    row = gj['rows'][0]
    assert len(gj['rows']) == 1
    assert row['debits'][0]['amount'] == Decimal('1000.00')
    assert row['credits'][0]['amount'] == Decimal('1000.00')
    assert row['explanation'] == 'Test adjustment'
    assert gj['total_debit'] == Decimal('1000.00') and gj['balanced'] is True


def test_draft_and_voided_excluded_from_totals(db_session, main_branch,
                                               cash_account, revenue_account):
    _entry(main_branch.id, cash_account, revenue_account, Decimal('500'), status='draft')
    _entry(main_branch.id, cash_account, revenue_account, Decimal('700'), status='cancelled')
    gj = build_general_journal(JournalEntry.query.all())
    assert gj['total_debit'] == Decimal('0.00')
    assert {r['is_draft'] for r in gj['rows']} == {True, False}
    assert {r['is_voided'] for r in gj['rows']} == {True, False}


def test_build_general_journal_xlsx_has_header_and_totals(db_session, main_branch,
                                                          cash_account, revenue_account):
    from openpyxl import load_workbook
    _entry(main_branch.id, cash_account, revenue_account, Decimal('1000'))
    gj = build_general_journal(JournalEntry.query.all())
    company = {'name': 'Acme', 'tin': '123', 'tin_branch': '', 'rdo': '050', 'address': 'X'}
    bio = build_general_journal_xlsx(gj, 'For June 2026', company, None, 'gj.xlsx')
    ws = load_workbook(bio).active
    assert 'ACME' in [c.value for c in ws['A']]
    assert 'GENERAL JOURNAL' in [c.value for c in ws['A']]
    assert 'TOTAL' in [c.value for c in ws['B']]
