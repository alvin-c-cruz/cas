import pytest
from io import BytesIO
from datetime import date
from openpyxl import load_workbook
from app.reports.statement_export import cash_flow_lines, build_cash_flow_xlsx

pytestmark = [pytest.mark.unit]

# Two-column (merged) indirect Cash Flow, same value in both columns for clarity.
CF = {
    'method': 'indirect', 'as_of': date(2026, 7, 31),
    'operating': {
        'net_income': {'mtd': 350.0, 'ytd': 350.0},
        'depreciation': {'mtd': 50.0, 'ytd': 50.0},
        'working_capital': [
            {'name': '(Increase)/decrease in Accounts Receivable', 'mtd_amount': -300.0, 'ytd_amount': -300.0},
            {'name': 'Increase/(decrease) in Accounts Payable', 'mtd_amount': 100.0, 'ytd_amount': 100.0},
        ],
        'total': {'mtd': 200.0, 'ytd': 200.0},
    },
    'investing': {'lines': [{'name': '(Acquisition)/disposal of Construction Equipment',
                             'mtd_amount': -500.0, 'ytd_amount': -500.0}],
                  'total': {'mtd': -500.0, 'ytd': -500.0}},
    'financing': {'lines': [{'name': 'Capital Stock', 'mtd_amount': 1000.0, 'ytd_amount': 1000.0}],
                  'total': {'mtd': 1000.0, 'ytd': 1000.0}},
    'net_change': {'mtd': 700.0, 'ytd': 700.0},
    'cash_begin': {'mtd': 0.0, 'ytd': 0.0}, 'cash_end': {'mtd': 700.0, 'ytd': 700.0},
    'is_reconciled': True, 'difference': {'mtd': 0.0, 'ytd': 0.0},
}


def test_lines_cover_all_sections():
    lines = cash_flow_lines(CF)
    labels = [ln['label'] for ln in lines]
    assert 'CASH FLOWS FROM OPERATING ACTIVITIES' in labels
    assert 'CASH FLOWS FROM INVESTING ACTIVITIES' in labels
    assert 'CASH FLOWS FROM FINANCING ACTIVITIES' in labels
    assert 'NET INCREASE/(DECREASE) IN CASH' in labels
    assert 'Cash at beginning of period' in labels
    assert 'Cash at end of period' in labels
    net = next(ln for ln in lines if ln['label'] == 'NET INCREASE/(DECREASE) IN CASH')
    assert net['mtd'] == 700.0 and net['ytd'] == 700.0 and net['rule'] == 'double_bottom'


def test_lines_carry_two_columns():
    lines = cash_flow_lines(CF)
    ni = next(ln for ln in lines if ln['label'] == 'Net Income (period)')
    assert ni['mtd'] == 350.0 and ni['ytd'] == 350.0


def test_depreciation_line_omitted_when_zero():
    cf = {**CF, 'operating': {**CF['operating'], 'depreciation': {'mtd': 0.0, 'ytd': 0.0}}}
    labels = [ln['label'] for ln in cash_flow_lines(cf)]
    assert not any('Depreciation' in lbl for lbl in labels)


def test_xlsx_has_two_amount_column_headers():
    xlsx = build_cash_flow_xlsx(CF, as_of_label='As of July 31, 2026',
                                company={}, branch_name=None, filename='cf.xlsx')
    raw = xlsx.data if hasattr(xlsx, 'data') else xlsx
    cells = [c.value for row in load_workbook(BytesIO(raw)).active.iter_rows()
             for c in row if c.value is not None]
    assert 'Jul 2026' in cells
    assert 'YTD 2026' in cells
    assert 'NET INCREASE/(DECREASE) IN CASH' in cells
