"""P6 Task 6 -- abnormal loss on the Cost of Production Report.

P5's statement gains three things, one per schedule:

    quantity schedule   the single "Unaccounted" line splits into Normal / Abnormal
    equivalent units    an abnormal-loss line, so the denominator is legible
    costs accounted for "Abnormal loss charged out", and the tie-out becomes
                        transferred + abnormal charged + ending WIP == pool

**The charged amount is read off the LEDGER, not recomputed.** P5's whole design
rule is that three of its five cost figures come from posted journal entries, which
is what makes "costs accounted for == costs to account for" a reconciliation of the
WIP control account rather than the report agreeing with itself. An abnormal figure
recomputed from the run record would be exempt from that check -- and the tie-out
would then stay green through exactly the drift it exists to catch.

Unit COUNTS are different and legitimately come from the run: no journal entry
carries a unit count, which is the same reason equivalent_units() is shared.
"""
from datetime import date
from decimal import Decimal

import pytest

from app import db
from app.bill_of_materials.models import BillOfMaterial, BillOfMaterialLine
from app.journal_entries.models import JournalEntry
from app.manufacturing_departments.models import ManufacturingDepartment
from app.production_runs.models import ProductionRun
from app.production_runs.service import close_run, issue_material, snapshot_materials
from app.products.models import Product
from app.reports.production_run_costing import generate_production_run_cost_report
from app.stock_adjustments.service import post_movement

pytestmark = [pytest.mark.integration, pytest.mark.reports, pytest.mark.production_runs]

_N = [0]


@pytest.fixture
def abnormal_loss_account(db_session, make_account):
    from app.settings import AppSettings
    make_account('7103')
    AppSettings.set_setting('abnormal_loss_account_code', '7103', updated_by='test')


def _closed_run(branch, actor, suffix, *, normal_loss_pct=None, completed='70',
                ending='20'):
    """1000.00 material + 450.00 conversion = P4/P5's 1450.00 pool.

    With a 3% expectation and 10 units lost: 3 normal, 7 abnormal, EU = 87,
    16.67 per EU, 1166.90 transferred and 116.69 charged out, leaving 166.41.
    """
    comp = Product(code=f'R6-C-{suffix}', name='Fresh Mango', track_inventory=True,
                   costing_method='moving_average', standard_cost=Decimal('5.00'),
                   is_active=True)
    out = Product(code=f'R6-O-{suffix}', name='Dried Mango', track_inventory=True,
                  costing_method='moving_average', is_active=True)
    db.session.add_all([comp, out]); db.session.commit()
    post_movement(comp, branch.id, 'opening', Decimal('10000'), Decimal('5.00'),
                  'stock_adjustment', 0, 'seed', actor, movement_date=date(2026, 1, 1))
    db.session.commit()
    bom = BillOfMaterial(product_id=out.id, manufacturing_mode='process',
                         normal_loss_pct=normal_loss_pct)
    bom.lines.append(BillOfMaterialLine(line_number=1, component_product_id=comp.id,
                                        quantity_per=Decimal('2')))
    db.session.add(bom); db.session.commit()
    dept = ManufacturingDepartment(branch_id=branch.id, code=f'S{suffix}', name='Dehydration')
    db.session.add(dept); db.session.commit()
    _N[0] += 1
    run = ProductionRun(run_number='R6%04d' % _N[0], bom_id=bom.id, department_id=dept.id,
                        branch_id=branch.id, period_start=date(2026, 8, 1),
                        period_end=date(2026, 8, 31), units_started=Decimal('100'),
                        conversion_cost=Decimal('450.00'),
                        units_completed_and_transferred=Decimal(completed),
                        units_ending_wip=Decimal(ending),
                        ending_wip_pct_complete=Decimal('50'))
    db.session.add(run); db.session.commit()
    snapshot_materials(run); db.session.commit()
    issue_material(run.materials[0], Decimal('200'), actor); db.session.commit()
    close_run(run, actor); db.session.commit()
    return run


def _login(client, user, branch):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id); s['_fresh'] = True
        s['selected_branch_id'] = branch.id


class TestTheQuantityScheduleSplitsTheLoss:
    def test_normal_and_abnormal_are_reported_separately(
            self, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        run = _closed_run(main_branch, accountant_user, 'A',
                          normal_loss_pct=Decimal('3.00'))
        rep = generate_production_run_cost_report(run.id, main_branch.id)
        assert rep['unaccounted_units'] == Decimal('10.0000'), "P5's total, unmoved"
        assert rep['normal_loss_units'] == Decimal('3.0000')
        assert rep['abnormal_loss_units'] == Decimal('7.0000')
        assert rep['normal_loss_units'] + rep['abnormal_loss_units'] == \
            rep['unaccounted_units'], 'the split must be exhaustive'

    def test_with_no_expectation_the_whole_loss_is_normal(
            self, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        """The backward-compatibility guarantee on the statement itself."""
        run = _closed_run(main_branch, accountant_user, 'B')
        rep = generate_production_run_cost_report(run.id, main_branch.id)
        assert rep['unaccounted_units'] == Decimal('10.0000')
        assert rep['normal_loss_units'] == Decimal('10.0000')
        assert rep['abnormal_loss_units'] == Decimal('0.0000')


class TestTheEquivalentUnitsScheduleIsLegible:
    def test_the_ending_wip_and_abnormal_terms_are_reported_separately(
            self, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        """The template used to derive the ending-WIP term by subtracting completed
        units from EU. With a third term in the denominator that subtraction silently
        folds abnormal loss into the ending-WIP line, so the generator now hands over
        both terms explicitly: 70 + 10 + 7 = 87."""
        run = _closed_run(main_branch, accountant_user, 'C',
                          normal_loss_pct=Decimal('3.00'))
        rep = generate_production_run_cost_report(run.id, main_branch.id)
        assert rep['ending_wip_equivalent_units'] == Decimal('10.0000')
        assert rep['abnormal_loss_units'] == Decimal('7.0000')
        assert rep['equivalent_units'] == Decimal('87.0000')
        assert (rep['units_completed_and_transferred']
                + rep['ending_wip_equivalent_units']
                + rep['abnormal_loss_units']) == rep['equivalent_units']


class TestCostsAccountedForAndTheTieOut:
    def test_the_charge_is_read_off_the_ledger_and_the_statement_reconciles(
            self, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        run = _closed_run(main_branch, accountant_user, 'D',
                          normal_loss_pct=Decimal('3.00'))
        rep = generate_production_run_cost_report(run.id, main_branch.id)
        assert rep['total_to_account_for'] == Decimal('1450.00')
        assert rep['transferred_out'] == Decimal('1166.90')
        assert rep['abnormal_loss_charged'] == Decimal('116.69')
        assert rep['ending_wip_cost'] == Decimal('166.41')
        assert rep['total_accounted_for'] == Decimal('1450.00')
        assert rep['difference'] == Decimal('0.00')
        assert rep['reconciles'] is True

    def test_the_charge_comes_from_the_JOURNAL_not_the_run_record(
            self, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        """The tie-out has to be able to go RED, or it proves nothing. Deleting the
        abnormal-loss posting while leaving every stored figure intact is the exact
        drift this statement exists to catch: a recomputed figure would sail through.
        """
        run = _closed_run(main_branch, accountant_user, 'E',
                          normal_loss_pct=Decimal('3.00'))
        je = JournalEntry.query.filter_by(entry_type='manufacturing_abnormal_loss',
                                          reference=run.run_number).one()
        db.session.delete(je); db.session.commit()

        rep = generate_production_run_cost_report(run.id, main_branch.id)
        assert rep['abnormal_loss_charged'] == Decimal('0.00'), 'the ledger no longer has it'
        assert rep['reconciles'] is False
        assert rep['difference'] == Decimal('116.69'), 'exactly the missing posting'

    def test_a_run_with_nothing_abnormal_reports_zero_and_still_reconciles(
            self, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        """P5's own figures, byte-identical, for an install that set no expectation."""
        run = _closed_run(main_branch, accountant_user, 'F', completed='80')
        rep = generate_production_run_cost_report(run.id, main_branch.id)
        assert rep['abnormal_loss_charged'] == Decimal('0.00')
        assert rep['transferred_out'] == Decimal('1288.80')
        assert rep['ending_wip_cost'] == Decimal('161.20')
        assert rep['reconciles'] is True


class TestTheThreeSurfaces:
    def test_the_screen_renders_the_split_and_the_charge(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        run = _closed_run(main_branch, accountant_user, 'G',
                          normal_loss_pct=Decimal('3.00'))
        _login(client, accountant_user, main_branch)
        r = client.get(f'/reports/production-run-cost/{run.id}')
        assert r.status_code == 200
        assert b'Normal loss' in r.data
        assert b'Abnormal loss' in r.data
        assert b'Abnormal loss charged out' in r.data
        assert b'116.69' in r.data
        assert b'166.41' in r.data
        # The equivalent-units schedule used to derive its ending-WIP term by
        # subtracting completed units from EU. That subtraction now yields 17.0000
        # (10 ending-WIP equivalents PLUS the 7 abnormal units), which would print a
        # denominator whose lines do not mean what they say while still summing to 87.
        assert b'10.0000' in r.data, 'the ending-WIP equivalents, on their own'
        assert b'17.0000' not in r.data, 'not the two terms silently folded together'
        # The reconciled banner enumerates which figures come from the ledger, and
        # abnormal loss is now the fourth. Found by the live browser gate, not by
        # pytest: nothing asserted the banner's CONTENT, only that the page rendered,
        # so it kept claiming three sources while reconciling four.
        assert b'abnormal loss are read from' in r.data, \
            'the banner must name every ledger-read figure it reconciles'

    def test_the_print_view_renders_them_too(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        """A print view that quietly omits a charge the screen shows is the surface
        an accountant actually files."""
        run = _closed_run(main_branch, accountant_user, 'H',
                          normal_loss_pct=Decimal('3.00'))
        _login(client, accountant_user, main_branch)
        r = client.get(f'/reports/production-run-cost/{run.id}/print')
        assert r.status_code == 200
        assert b'Abnormal loss charged out' in r.data
        assert b'116.69' in r.data

    def test_the_excel_export_writes_the_new_rows_as_NUMBERS(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts,
            abnormal_loss_account):
        """Cell TYPES, not just values -- D5's defect was handing format_value() a
        float, which str()s into TEXT that merely looks like a number and silently
        breaks every SUM the accountant writes over it. Decimals stay Decimals."""
        from io import BytesIO
        from openpyxl import load_workbook

        run = _closed_run(main_branch, accountant_user, 'I',
                          normal_loss_pct=Decimal('3.00'))
        _login(client, accountant_user, main_branch)
        r = client.get(f'/reports/production-run-cost/{run.id}/export/excel')
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.data))
        ws = wb.active
        cells = {}
        for row in ws.iter_rows():
            label = row[0].value
            if label:
                cells[str(label).strip()] = row[1]
        assert cells['Normal loss'].value == 3
        assert isinstance(cells['Normal loss'].value, (int, float))
        assert cells['Abnormal loss'].value == 7
        assert isinstance(cells['Abnormal loss'].value, (int, float))
        assert cells['Abnormal loss charged out'].value == 116.69
        assert isinstance(cells['Abnormal loss charged out'].value, (int, float)), \
            'a float here would be TEXT in the sheet'
