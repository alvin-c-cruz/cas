"""P5 Tasks 3-5 -- the screen, print and Excel routes.

Mirrors D5's route trio (screen / print / export) and its authorization: both the
outer module gate and the view's own accountant_or_above check, tested separately
because the outer one stops an ungranted user first and would otherwise leave the
inner one with zero coverage (the P1 lesson, which recurred in P4).

The Excel test reads the returned workbook's CELL TYPES back. D5's plan told the
implementer to cast float(...), which export_to_excel's format_value() sends
through str() into a TEXT cell that only looks like a number -- and the plan's
specified "200 + Content-Type" test passes against BOTH versions, which is exactly
why that defect reached a browser pass.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest

from app import db
from app.bill_of_materials.models import BillOfMaterial, BillOfMaterialLine
from app.manufacturing_departments.models import ManufacturingDepartment
from app.production_runs.models import ProductionRun
from app.production_runs.service import close_run, issue_material, snapshot_materials
from app.products.models import Product
from app.stock_adjustments.service import post_movement

pytestmark = [pytest.mark.integration, pytest.mark.reports, pytest.mark.production_runs]

_N = [0]


def _enable(db_session):
    from app.settings import AppSettings
    from app.utils.cache_helpers import clear_module_config_cache
    for k in ('bill_of_materials', 'production_runs'):
        AppSettings.set_setting(f'module_enabled:{k}', '1')
    db_session.commit(); clear_module_config_cache()


def _login(client, user, branch):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id); s['_fresh'] = True
        s['selected_branch_id'] = branch.id


def _run(branch, actor, suffix, close=True, **run_kw):
    comp = Product(code=f'RT-C-{suffix}', name='Fresh Mango', track_inventory=True,
                   costing_method='moving_average', standard_cost=Decimal('5.00'),
                   is_active=True)
    out = Product(code=f'RT-O-{suffix}', name='Dried Mango', track_inventory=True,
                  costing_method='moving_average', is_active=True)
    db.session.add_all([comp, out]); db.session.commit()
    post_movement(comp, branch.id, 'opening', Decimal('10000'), Decimal('5.00'),
                  'stock_adjustment', 0, 'seed', actor, movement_date=date(2026, 1, 1))
    db.session.commit()
    bom = BillOfMaterial(product_id=out.id, manufacturing_mode='process')
    bom.lines.append(BillOfMaterialLine(line_number=1, component_product_id=comp.id,
                                        quantity_per=Decimal('2')))
    db.session.add(bom); db.session.commit()
    dept = ManufacturingDepartment(branch_id=branch.id, code=f'T{suffix}', name='Dehydration')
    db.session.add(dept); db.session.commit()
    _N[0] += 1
    kw = dict(conversion_cost=Decimal('450.00'),
              units_completed_and_transferred=Decimal('80'),
              units_ending_wip=Decimal('20'),
              ending_wip_pct_complete=Decimal('50'))
    kw.update(run_kw)
    run = ProductionRun(run_number='RT%04d' % _N[0], bom_id=bom.id, department_id=dept.id,
                        branch_id=branch.id, period_start=date(2026, 8, 1),
                        period_end=date(2026, 8, 31), units_started=Decimal('100'), **kw)
    db.session.add(run); db.session.commit()
    snapshot_materials(run); db.session.commit()
    issue_material(run.materials[0], Decimal('200'), actor); db.session.commit()
    if close:
        close_run(run, actor); db.session.commit()
    return run


class TestScreen:
    def test_renders_all_three_schedules_and_the_figures(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'A')
        _login(client, accountant_user, main_branch)
        resp = client.get(f'/reports/production-run-cost/{run.id}')
        assert resp.status_code == 200
        body = resp.data.decode('utf-8')
        for heading in ('Quantity schedule', 'Equivalent units', 'Costs to account for',
                        'Costs accounted for'):
            assert heading.lower() in body.lower(), heading
        for figure in ('1000.00', '450.00', '1450.00', '1288.80', '161.20', '16.11', '90.0000'):
            assert figure in body, figure

    def test_shows_the_reconciled_state(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'B')
        _login(client, accountant_user, main_branch)
        body = client.get(f'/reports/production-run-cost/{run.id}').data.decode('utf-8')
        assert 'Reconciled' in body
        assert 'does not reconcile' not in body.lower()

    def test_shows_the_FAILING_state_with_the_difference_named(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        """The state the report exists for. It must say so loudly and name the gap."""
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'C')
        run.ending_wip_cost = Decimal('150.00')      # was 161.20
        db.session.commit()
        _login(client, accountant_user, main_branch)
        body = client.get(f'/reports/production-run-cost/{run.id}').data.decode('utf-8')
        assert 'does not reconcile' in body.lower()
        assert '11.20' in body, 'the difference must be shown, not just flagged'

    def test_a_negative_unaccounted_figure_is_flagged(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'D',
                   units_completed_and_transferred=Decimal('95'),
                   units_ending_wip=Decimal('40'))
        _login(client, accountant_user, main_branch)
        body = client.get(f'/reports/production-run-cost/{run.id}').data.decode('utf-8')
        assert 'unaccounted' in body.lower()
        assert 'data error' in body.lower(), 'a negative figure needs explaining, not just colour'

    def test_an_open_run_is_refused(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'E', close=False)
        _login(client, accountant_user, main_branch)
        resp = client.get(f'/reports/production-run-cost/{run.id}', follow_redirects=True)
        assert b'closed' in resp.data.lower()

    def test_another_branchs_run_is_refused(
            self, client, db_session, main_branch, branch_manila, accountant_user,
            wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'F')
        accountant_user.set_branches([main_branch, branch_manila]); db_session.commit()
        _login(client, accountant_user, branch_manila)
        resp = client.get(f'/reports/production-run-cost/{run.id}', follow_redirects=True)
        assert b'does not exist in this branch' in resp.data


class TestPrint:
    def test_renders_with_company_and_branch_headings(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'G')
        _login(client, accountant_user, main_branch)
        resp = client.get(f'/reports/production-run-cost/{run.id}/print')
        assert resp.status_code == 200
        body = resp.data.decode('utf-8')
        assert 'COST OF PRODUCTION REPORT' in body.upper()
        assert '1288.80' in body


class TestExcel:
    def _workbook(self, resp):
        from openpyxl import load_workbook
        return load_workbook(BytesIO(resp.data))

    def test_exports_and_is_an_xlsx(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'H')
        _login(client, accountant_user, main_branch)
        resp = client.get(f'/reports/production-run-cost/{run.id}/export/excel')
        assert resp.status_code == 200
        assert 'spreadsheet' in resp.headers['Content-Type']

    def test_money_cells_are_NUMBERS_not_text(
            self, client, db_session, main_branch, accountant_user, wo_control_accounts):
        """Reads the workbook back. A 200 + Content-Type assertion passes against a
        version that writes every figure as a string, which is how the same defect
        shipped in D5."""
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'I')
        _login(client, accountant_user, main_branch)
        resp = client.get(f'/reports/production-run-cost/{run.id}/export/excel')
        ws = self._workbook(resp).active

        values = {}
        for row in ws.iter_rows():
            cells = [c for c in row]
            if len(cells) >= 2 and isinstance(cells[0].value, str) and cells[1].value is not None:
                values[cells[0].value.strip().lower()] = cells[1]

        money_rows = [k for k in values if 'material added' in k or 'transferred out' in k
                      or 'total' in k]
        assert money_rows, f'no money rows found; sheet labels were {list(values)}'
        for k in money_rows:
            cell = values[k]
            assert not isinstance(cell.value, str), \
                f'{k!r} exported as TEXT ({cell.value!r}) -- a number that only looks like one'
            assert isinstance(cell.value, (int, float)), f'{k!r} is {type(cell.value)}'


class TestAuthorization:
    def test_ungranted_staff_stopped_by_the_outer_module_gate(
            self, client, db_session, main_branch, staff_user, accountant_user,
            wo_control_accounts):
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'J')
        staff_user.set_branches([main_branch]); db_session.commit()
        _login(client, staff_user, main_branch)
        resp = client.get(f'/reports/production-run-cost/{run.id}', follow_redirects=True)
        assert resp.status_code in (200, 403)
        assert b'1288.80' not in resp.data, 'a staff user must not see the figures'

    def test_granted_staff_still_refused_by_the_views_own_guard(
            self, client, db_session, main_branch, staff_user, accountant_user,
            wo_control_accounts):
        """Past the outer gate, so this is the only test that can see the view's own
        accountant_or_above check."""
        _enable(db_session)
        run = _run(main_branch, accountant_user, 'K')
        staff_user.set_branches([main_branch])
        staff_user.set_book_permissions({'production_runs': True, 'bill_of_materials': True,
                                         'reports': True})
        db_session.commit()
        _login(client, staff_user, main_branch)
        resp = client.get(f'/reports/production-run-cost/{run.id}', follow_redirects=True)
        assert b'1288.80' not in resp.data
